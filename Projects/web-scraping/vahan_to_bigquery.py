import re, sys, logging
from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)])
log = logging.getLogger(__name__)

MONTH_ORDER = ['JAN','FEB','MAR','APR','MAY','JUN','JUL','AUG','SEP','OCT','NOV','DEC']

OEM_MAP = {
    "MARUTI SUZUKI INDIA LTD":"Maruti","MARUTI UDYOG LTD":"Maruti",
    "HYUNDAI MOTOR INDIA LTD":"Hyundai",
    "TATA MOTORS LTD":"Tata","TATA MOTORS PASSENGER VEHICLES LTD":"Tata",
    "TATA PASSENGER ELECTRIC MOBILITY LTD":"Tata","TATA ADVANCED SYSTEMS LTD":"Tata",
    "MAHINDRA & MAHINDRA LIMITED":"Mahindra","MAHINDRA & MAHINDRA LIMITED (SWARAJ DIVISION)":"Swaraj",
    "MAHINDRA & MAHINDRA LIMITED (TRACTOR)":"Mahindra",
    "MAHINDRA ELECTRIC AUTOMOBILE LTD":"Mahindra","MAHINDRA ELECTRIC MOBILITY LIMITED":"Mahindra",
    "MAHINDRA LAST MILE MOBILITY LTD":"Mahindra-LMM","MAHINDRA DEFENCE SYSTEMS LTD":"Mahindra",
    "TOYOTA KIRLOSKAR MOTOR PVT LTD":"Toyota","TOYOTA MATERIAL HANDLING INDIA PVT LTD":"Toyota",
    "HONDA CARS INDIA LTD":"Honda","HONDA MOTORCYCLE AND SCOOTER INDIA (P) LTD":"Honda",
    "KIA INDIA PRIVATE LIMITED":"Kia","SKODA AUTO VOLKSWAGEN INDIA PVT LTD":"Volkswagen",
    "SKODA AUTO AS":"Skoda","VOLKSWAGEN AG":"Volkswagen",
    "RENAULT INDIA PVT LTD":"Renault","NISSAN MOTOR INDIA PVT LTD":"Nissan",
    "JSW MG MOTOR INDIA PVT LTD":"MG","BMW INDIA PVT LTD":"BMW",
    "MERCEDES-BENZ INDIA PVT LTD":"Mercedes-Benz","MERCEDES -BENZ AG":"Mercedes-Benz",
    "AUDI AG":"Audi","PORSCHE AG GERMANY":"Porsche",
    "JAGUAR LAND ROVER INDIA LIMITED":"JLR","FORD INDIA PVT LTD":"Ford",
    "ISUZU MOTORS INDIA PVT LTD":"Isuzu","SML ISUZU LTD":"SML Isuzu",
    "FERRARI SPA (IMPORTER:NAVNIT MOTORS PVT LTD)":"Ferrari",
    "FERRARI SPA (IMPORTER:SELECT CARS P LTD)":"Ferrari",
    "ROLLS-ROYCE MOTOR CARS(IMPORTER:SELECT CARS P LTD)":"Rolls-Royce",
    "Rolls Royce Motor (Importer: KUN Motor)":"Rolls-Royce",
    "BENTLEY MOTORS LTD (SAVWIPL)":"Bentley","AUTOMOBILI LAMBORGHINI S.P.A":"Lamborghini",
    "ASTON MARTIN LAGONDA LTD":"Aston Martin","TESLA INDIA MOTORS AND ENERGY PVT LTD":"Tesla",
    "VINFAST AUTO INDIA PVT LTD":"VinFast",
    "HERO MOTOCORP LTD":"Hero","HERO HONDA MOTORS LTD":"Hero",
    "HERO ELECTRIC VEHICLES PVT. LTD":"Hero Electric",
    "BAJAJ AUTO LTD":"Bajaj","TVS MOTOR COMPANY LTD":"TVS",
    "ROYAL-ENFIELD (UNIT OF EICHER LTD)":"Royal Enfield",
    "INDIA YAMAHA MOTOR PVT LTD":"Yamaha","SUZUKI MOTORCYCLE INDIA PVT LTD":"Suzuki",
    "INDIA KAWASAKI MOTORS PVT LTD":"Kawasaki","TRIUMPH MOTORCYCLES (INDIA) PVT LTD":"Triumph",
    "DUCATI INDIA PVT LTD":"Ducati","HARLEY DAVIDSON (IMPORTER: HERO MOTOCORP)":"Harley-Davidson",
    "CLASSIC LEGENDS PVT LTD":"Classic Legends","ATHER ENERGY LTD":"Ather",
    "OLA ELECTRIC TECHNOLOGIES PVT LTD":"Ola Electric","REVOLT INTELLICORP PVT LTD":"Revolt",
    "ULTRAVIOLETTE AUTOMOTIVE PVT LTD":"Ultraviolette",
    "ASHOK LEYLAND LTD":"Ashok Leyland","VE COMMERCIAL VEHICLES LTD":"Volvo Eicher",
    "VE COMMERCIAL VEHICLES LTD (VOLVO BUSES DIVISION)":"Volvo Eicher",
    "DAIMLER INDIA COMMERCIAL VEHICLES PVT. LTD":"Daimler",
    "SCANIA COMMERCIAL VEHICLES INDIA PVT LIMITED":"Scania",
    "VOLVO GROUP INDIA PVT LTD":"Volvo","VOLVO AUTO INDIA PVT LTD":"Volvo",
    "OLECTRA GREENTECH LTD":"Olectra","JBM AUTO LIMITED":"JBM",
    "JBM ELECTRIC VEHICLES PVT LTD":"JBM","FORCE MOTORS LIMITED":"Force Motors",
    "PIAGGIO VEHICLES PVT LTD":"Piaggio","ATUL AUTO LTD":"Atul Auto",
    "EULER MOTORS PVT LTD":"Euler Motors","OMEGA SEIKI PVT LTD":"Omega Seiki",
    "INTERNATIONAL TRACTORS LIMITED":"Sonalika",
    "SONALIKA INTERNATIONAL TRACTORS LIMITED":"Sonalika",
    "JOHN DEERE INDIA PVT LTD(TRACTOR DEVISION)":"John Deere",
    "ESCORTS KUBOTA LIMITED (AGRI MACHINERY GROUP)":"Escorts Kubota",
    "ESCORTS KUBOTA LIMITED (CONSTRUCTION EQUIPMENT)":"Escorts Kubota",
    "ESCORTS LTD":"Escorts","CNH INDUSTRIAL (INDIA) PVT LTD":"CNH",
    "TAFE LIMITED":"TAFE","KUBOTA AGRICULTURAL MACHINERY INDIA PVT.LTD.":"Kubota",
    "JCB INDIA LIMITED":"JCB","KOMATSU INDIA PRIVATE LIMITED":"Komatsu",
    "CATERPILLAR INDIA PRIVATE LIMITED":"Caterpillar",
    "SANY HEAVY INDUSTRY INDIA PVT LTD":"Sany","SANY INDIA":"Sany",
    "XCMG CONSTRUCTION MACHINERY CO. LTD.":"XCMG",
    "VOLVO CE INDIA PRIVATE LIMITED":"Volvo CE",
    "HD HYUNDAI CONSTRUCTION EQUIPMENT INDIA PVT LTD":"HD Hyundai CE",
    "AMMANN APOLLO INDIA PVT LTD":"Ammann Apollo",
    "GREAVES ELECTRIC MOBILITY PVT LTD":"Greaves Electric",
    "BYD INDIA PRIVATE LIMITED":"BYD","AJAX ENGINEERING LTD":"Ajax",
    "ACTION CONSTRUCTION EQUIPMENT LTD.":"ACE",
    "GREAVES ELECTRIC MOBILITY PVT LTD":"Greaves Electric",
    "AMPERE VEHICLES PRIVATE LIMITED":"Ampere",
    "KINETIC GREEN ENERGY & POWER SOLUTIONS LTD":"Kinetic Green",
    "SKS TRADE INDIA PVT LTD":"SKS",
}

MASTER_FUEL_MAP = {
    "PETROL":"Petrol","DIESEL":"Diesel","CNG ONLY":"CNG","PETROL/CNG":"CNG",
    "ELECTRIC(BOV)":"Electric","PURE EV":"Electric","STRONG HYBRID EV":"Strong Hybrid EV",
    "PETROL/HYBRID":"Petrol - Mild Hybrid","DIESEL/HYBRID":"Diesel - Mild Hybrid",
    "PETROL(E20)/HYBRID":"Petrol - Mild Hybrid","PETROL/ETHANOL":"Petrol - Ethanol",
    "FUEL CELL HYDROGEN":"Hydrogen","LPG ONLY":"LPG","PETROL/LPG":"LPG",
    "DIESEL/CNG":"CNG","SOLAR":"Solar",
}

def get_oem(maker):
    if not maker or pd.isna(maker): return ""
    m = str(maker).strip()
    if m in OEM_MAP: return OEM_MAP[m]
    ml = m.upper()
    for key, val in OEM_MAP.items():
        if key.upper() in ml: return val
    words = re.sub(r'\(.*?\)', '', m).strip().split()
    return " ".join(words[:2]).title() if words else m

def get_master_fuel(fuel):
    if not fuel or pd.isna(fuel): return ""
    return MASTER_FUEL_MAP.get(str(fuel).strip().upper(), str(fuel).strip().title())

def get_fy(tp):
    y, m = int(str(tp)[:4]), int(str(tp)[5:7])
    fy = y+1 if m >= 4 else y
    return f"F{str(fy)[2:]}"

def get_quarter(tp):
    y, m = int(str(tp)[:4]), int(str(tp)[5:7])
    fy = y+1 if m >= 4 else y
    q = ((m-4) % 12) // 3 + 1
    return f"F{str(fy)[2:]}Q{q}"

# ── Read wide-format Excel ────────────────────────────────────────────────────
def read_wide(path):
    df = pd.read_excel(path, header=[0,1])
    df.columns = pd.MultiIndex.from_tuples(df.columns)

    id_cols  = [c for g,c in df.columns if g=='IDENTIFICATION']
    grp_map  = {}
    for g,c in df.columns:
        if g != 'IDENTIFICATION':
            grp_map.setdefault(g, []).append(c)

    # Flatten for easy access
    df.columns = [f"{g}||{c}" for g,c in df.columns]
    id_flat  = [f"IDENTIFICATION||{c}" for c in id_cols]

    return df, id_flat, grp_map

# ── Melt one group into long rows ─────────────────────────────────────────────
def melt_group(df, id_flat, group_name, cat_cols, breakdown_type, created_dt):
    rows = []
    for _, row in df.iterrows():
        tp    = str(row['IDENTIFICATION||Period'])   # e.g. '2025-01'
        month_val = str(row['IDENTIFICATION||Month']) # e.g. 'JAN'

        base = {
            'financial_year': get_fy(tp),
            'l_quarter':      get_quarter(tp),
            'month':          f"{tp}-01",
            'state_code':     row['IDENTIFICATION||State Code'],
            'state_name':     row['IDENTIFICATION||State'],
            'rto_code':       row['IDENTIFICATION||RTO Code'],
            'rto_name':       row['IDENTIFICATION||RTO Name'],
            'breakdown_type': breakdown_type,
            'maker':          '',
            'oem':            '',
            'vehicle_category':'',
            'vehicle_class':  '',
            'master_fuel':    '',
            'fuel':           '',
            'norms':          '',
            'mdp_created_dt': created_dt,
        }

        for cat in cat_cols:
            col_key = f"{group_name}||{cat}"
            if col_key not in df.columns:
                continue
            val = row[col_key]
            reg = 0
            try: reg = int(float(val)) if pd.notna(val) else 0
            except: reg = 0
            if reg <= 0:
                continue

            r = base.copy()
            r['registrations'] = reg

            if breakdown_type == 'fuel':
                r['fuel'] = cat
                r['master_fuel'] = get_master_fuel(cat)
            elif breakdown_type == 'maker':
                r['maker'] = cat
                r['oem']   = get_oem(cat)
            elif breakdown_type == 'vehicle_category':
                r['vehicle_category'] = cat
            elif breakdown_type == 'vehicle_class':
                r['vehicle_class'] = cat
            elif breakdown_type == 'norms':
                r['norms'] = cat

            rows.append(r)
    return rows

# ── Excel writer ──────────────────────────────────────────────────────────────
SHEET_COLORS = {
    'Fuel':             '833C00',
    'Maker':            '5C3317',
    'Vehicle Category': '375623',
    'Vehicle Class':    '4B2F8A',
    'Norms':            '1F4E79',
    'All Data':         '1F3864',
}

def write_excel(sheets_data: dict, output_path: Path):
    wb = Workbook()
    wb.remove(wb.active)

    COLS = ['financial_year','l_quarter','month','state_code','state_name',
            'rto_code','rto_name','breakdown_type','maker','oem',
            'vehicle_category','vehicle_class','master_fuel','fuel',
            'norms','registrations','mdp_created_dt']
    COL_W = {
        'financial_year':13,'l_quarter':11,'month':12,'state_code':10,
        'state_name':14,'rto_code':9,'rto_name':32,'breakdown_type':16,
        'maker':36,'oem':18,'vehicle_category':22,'vehicle_class':28,
        'master_fuel':20,'fuel':20,'norms':22,'registrations':14,'mdp_created_dt':22,
    }

    for sheet_name, df_sheet in sheets_data.items():
        ws = wb.create_sheet(sheet_name)
        color = SHEET_COLORS.get(sheet_name, '1F3864')

        s = Side(style='thin', color='C0C0C0')
        BD = Border(left=s, right=s, top=s, bottom=s)
        HDR_FILL = PatternFill('solid', fgColor=color)
        HDR_FONT = Font(bold=True, color='FFFFFF', name='Calibri', size=10)
        ALT_FILL = PatternFill('solid', fgColor='F2F7FD')
        DATA_FONT = Font(name='Calibri', size=9)

        # Header
        for ci, col in enumerate(COLS, 1):
            c = ws.cell(1, ci, col)
            c.font = HDR_FONT; c.fill = HDR_FILL; c.border = BD
            c.alignment = Alignment(horizontal='center', vertical='center')
            ws.column_dimensions[get_column_letter(ci)].width = COL_W.get(col, 14)
        ws.row_dimensions[1].height = 18

        # Data
        for ri, row_vals in enumerate(df_sheet[COLS].itertuples(index=False), 2):
            alt = ri % 2 == 0
            for ci, val in enumerate(row_vals, 1):
                if hasattr(val, 'item'): val = val.item()
                c = ws.cell(ri, ci, val)
                c.font = DATA_FONT; c.border = BD
                col_name = COLS[ci-1]
                c.alignment = Alignment(
                    horizontal='center' if col_name in ('registrations','rto_code','state_code','financial_year','l_quarter','month') else 'left',
                    vertical='center')
                if col_name == 'registrations':
                    c.number_format = '#,##0'
                if alt: c.fill = ALT_FILL
            ws.row_dimensions[ri].height = 13

        ws.freeze_panes = 'A2'
        ws.sheet_properties.tabColor = color
        log.info(f"  Sheet '{sheet_name}': {len(df_sheet):,} rows")

    wb.save(output_path)
    log.info(f"Saved: {output_path}  ({round(output_path.stat().st_size/1024)} KB)")

# ── Main ──────────────────────────────────────────────────────────────────────
def convert(input_path, output_path):
    log.info(f"Loading: {input_path}")
    df, id_flat, grp_map = read_wide(input_path)

    created_dt = datetime.now().strftime('%Y-%m-%d %H:%M:%S UTC')
    log.info(f"Rows: {len(df)} | Groups: {list(grp_map.keys())}")

    GROUP_TO_BT = {
        'BY FUEL TYPE':         ('fuel',             'Fuel'),
        'BY VEHICLE CATEGORY':  ('vehicle_category', 'Vehicle Category'),
        'BY VEHICLE CLASS':     ('vehicle_class',    'Vehicle Class'),
        'BY EMISSION NORMS':    ('norms',            'Norms'),
        'BY MAKER / BRAND':     ('maker',            'Maker'),
    }

    sheets_data = {}
    all_rows = []

    for grp_name, (bt, sheet_name) in GROUP_TO_BT.items():
        if grp_name not in grp_map:
            continue
        rows = melt_group(df, id_flat, grp_name, grp_map[grp_name], bt, created_dt)
        if rows:
            sheet_df = pd.DataFrame(rows)
            sheets_data[sheet_name] = sheet_df
            all_rows.extend(rows)
            log.info(f"  {sheet_name}: {len(rows):,} rows")

    if all_rows:
        all_df = pd.DataFrame(all_rows)
        all_df = all_df.sort_values(['state_code','rto_code','month','breakdown_type']).reset_index(drop=True)
        sheets_data['All Data'] = all_df

    # Put All Data first
    ordered = {'All Data': sheets_data.pop('All Data')} if 'All Data' in sheets_data else {}
    ordered.update(sheets_data)

    write_excel(ordered, output_path)

if __name__ == '__main__':
    import argparse
    p = argparse.ArgumentParser()
    p.add_argument('--input', default=None)
    p.add_argument('--output', default=None)
    args = p.parse_args()

    data_dir = Path('vahan_data')
    # Default: look for the most recent formatted file
    if args.input:
        inp = Path(args.input)
    else:
        candidates = sorted(data_dir.glob('VAHAN_*_formatted.xlsx'), key=lambda x: x.stat().st_mtime, reverse=True)
        if not candidates:
            log.error("No formatted file found in vahan_data/. Run vahan_format.py first.")
            sys.exit(1)
        inp = candidates[0]
        log.info(f"Auto-selected: {inp.name}")

    out = Path(args.output) if args.output else \
          data_dir / f"VAHAN_bigquery_{datetime.now().strftime('%Y%m%d')}.xlsx"

    convert(inp, out)
