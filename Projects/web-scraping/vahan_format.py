# # -*- coding: utf-8 -*-
# """
# VAHAN Data Formatter — single wide sheet output.
# Each row = RTO × Month. Columns = ID + all fuel types + vehicle categories +
# vehicle classes + emission norms + makers + TOTAL.

# USAGE:
#     python vahan_format.py
#     python vahan_format.py --input data.xlsx --output report.xlsx
#     python vahan_format.py --states DL UP
# """
# import re, sys, argparse, logging
# from pathlib import Path
# from datetime import datetime

# import pandas as pd
# from openpyxl import Workbook
# from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
# from openpyxl.utils import get_column_letter

# logging.basicConfig(level=logging.INFO,
#     format="%(asctime)s [%(levelname)s] %(message)s",
#     handlers=[logging.StreamHandler(sys.stdout)])
# log = logging.getLogger(__name__)

# MONTH_ORDER = ['JAN','FEB','MAR','APR','MAY','JUN','JUL','AUG','SEP','OCT','NOV','DEC']
# ID_COLS = ['Period','Month','Year','State Code','State','RTO Code','RTO Name']

# BREAKDOWNS = [
#     ('fuel',             'fuel',             'BY FUEL TYPE'),
#     ('vehicle_category', 'vehicle_category', 'BY VEHICLE CATEGORY'),
#     ('vehicle_class',    'vehicle_class',    'BY VEHICLE CLASS'),
#     ('norms',            'norms',            'BY EMISSION NORMS'),
#     ('maker',            'maker',            'BY MAKER / BRAND'),
# ]
# GROUP_LABELS = {
#     'info':'IDENTIFICATION','fuel':'BY FUEL TYPE','vehicle_category':'BY VEHICLE CATEGORY',
#     'vehicle_class':'BY VEHICLE CLASS','norms':'BY EMISSION NORMS',
#     'maker':'BY MAKER / BRAND','total':'TOTAL',
# }
# GROUP_STYLE = {
#     'info':             ('1F3864','D6E4F7'),
#     'fuel':             ('833C00','FCE4D6'),
#     'vehicle_category': ('375623','E2EFDA'),
#     'vehicle_class':    ('4B2F8A','EAE0F7'),
#     'norms':            ('1F4E79','DDEBF7'),
#     'maker':            ('5C3317','F2E0D0'),
#     'total':            ('404040','C6EFCE'),
# }

# # ── Style helpers ─────────────────────────────────────────────────────────────
# def hf(sz=10):              return Font(bold=True,color='FFFFFF',name='Calibri',size=sz)
# def df_(sz=9,bold=False):   return Font(name='Calibri',size=sz,bold=bold)
# def fl(c):                  return PatternFill('solid',fgColor=c)
# def bd():
#     s=Side(style='thin',color='C0C0C0')
#     return Border(left=s,right=s,top=s,bottom=s)
# def ctr(): return Alignment(horizontal='center',vertical='center',wrap_text=False)
# def lft(): return Alignment(horizontal='left',  vertical='center',wrap_text=False)

# def auto_w(ws,min_w=6,max_w=28):
#     for col in ws.columns:
#         w=max(min_w,min(max_w,max((len(str(c.value)) if c.value is not None else 0) for c in col)+2))
#         ws.column_dimensions[get_column_letter(col[0].column)].width=w

# def clean_rto(name):
#     m=re.match(r'^(.+?)\s*-\s*([A-Z]+\d+)\s*\(.*\)$',str(name).strip())
#     return f"{m.group(2)} \u2013 {m.group(1).strip().title()}" if m else str(name).strip()

# def clean_state(name):
#     m=re.match(r'^(.+?)\s*\(',str(name).strip())
#     return m.group(1).strip() if m else str(name).strip()

# # ── Core builder ──────────────────────────────────────────────────────────────
# def build_wide_sheet(df: pd.DataFrame, output_path: Path):
#     df = df.copy()

#     # Base grid: one row per RTO × month
#     base = (df[['time_period','month','year','state_code','state_clean','rto_code','rto_clean']]
#             .drop_duplicates()
#             .copy()
#             .rename(columns={'time_period':'Period','month':'Month','year':'Year',
#                              'state_code':'State Code','state_clean':'State',
#                              'rto_code':'RTO Code','rto_clean':'RTO Name'}))
#     base = base.assign(_mn=base['Month'].map({m:i for i,m in enumerate(MONTH_ORDER)}))
#     base = base.sort_values(['RTO Code','Year','_mn']).drop(columns='_mn').reset_index(drop=True)

#     # Pivot each breakdown and merge onto base
#     group_cols: dict[str, list] = {}
#     wide = base.copy()

#     for bt, cat_col, _ in BREAKDOWNS:
#         sub = df[df['breakdown_type'] == bt].copy()
#         if sub.empty:
#             continue
#         # Drop rows where the category value is purely numeric (scraper artefact —
#         # these are RTO codes that leaked into the category column)
#         sub = sub[sub[cat_col].astype(str).str.strip()
#                   .str.match(r'^[A-Za-z]').fillna(False)].copy()
#         if sub.empty:
#             continue
#         piv = (sub.pivot_table(
#                     index=['time_period','rto_code'],
#                     columns=cat_col,
#                     values='registrations_count',
#                     aggfunc='sum',
#                     fill_value=0)
#                .reset_index())
#         piv.columns.name = None
#         piv = piv.rename(columns={'time_period':'Period','rto_code':'RTO Code'})
#         cat_names = [c for c in piv.columns if c not in ('Period','RTO Code')]
#         group_cols[bt] = cat_names
#         wide = wide.merge(piv, on=['Period','RTO Code'], how='left')

#     # Fill NaN → 0 for all category columns
#     all_cat_cols = [c for cols in group_cols.values() for c in cols]
#     # Fix all numeric columns in one shot using assign — avoids pandas CoW warnings
#     num_updates = {c: pd.to_numeric(wide[c], errors='coerce').fillna(0).astype(int)
#                    for c in all_cat_cols if c in wide.columns}
#     wide = wide.assign(**num_updates)

#     # TOTAL = sum of fuel cols (or vehicle_category as fallback)
#     total_src = [c for c in (group_cols.get('fuel') or group_cols.get('vehicle_category') or [])
#                  if c in wide.columns]
#     wide = wide.assign(**{'TOTAL REGISTRATIONS': wide[total_src].sum(axis=1) if total_src else 0})
#     group_cols['total'] = ['TOTAL REGISTRATIONS']

#     # Build final column list — only include cols that actually exist in wide
#     final_cols = [c for c in ID_COLS if c in wide.columns]
#     for bt, _, _ in BREAKDOWNS:
#         final_cols += [c for c in group_cols.get(bt, []) if c in wide.columns]
#     final_cols += ['TOTAL REGISTRATIONS']

#     # Map each column to its group
#     col_group = {c: 'info' for c in ID_COLS}
#     for bt, _, _ in BREAKDOWNS:
#         for c in group_cols.get(bt, []):
#             col_group[c] = bt
#     col_group['TOTAL REGISTRATIONS'] = 'total'

#     # Safety: use reindex to guarantee all final_cols exist (fill missing with 0)
#     wide = wide.reindex(columns=final_cols, fill_value=0)

#     log.info(f"  Wide table: {len(wide)} rows x {len(final_cols)} columns")
#     for bt, _, label in BREAKDOWNS:
#         n = len([c for c in group_cols.get(bt, []) if c in wide.columns])
#         if n: log.info(f"    {label:28}: {n} cols")

#     # ── Write Excel ────────────────────────────────────────────────────────────
#     wb = Workbook()
#     ws = wb.active
#     ws.title = 'Vehicle Registration Data'

#     # Find column index ranges per group
#     grp_ranges: dict[str, tuple] = {}
#     for ci, col in enumerate(final_cols, 1):
#         grp = col_group.get(col, 'info')
#         grp_ranges[grp] = (grp_ranges[grp][0], ci) if grp in grp_ranges else (ci, ci)

#     # Row 1: merged group header cells
#     for grp, (s, e) in grp_ranges.items():
#         bg, _ = GROUP_STYLE[grp]
#         sl, el = get_column_letter(s), get_column_letter(e)
#         if s != e:
#             ws.merge_cells(f'{sl}1:{el}1')
#         cell = ws.cell(1, s, GROUP_LABELS.get(grp, grp.upper()))
#         cell.font=hf(10); cell.fill=fl(bg); cell.alignment=ctr(); cell.border=bd()
#         for ci2 in range(s+1, e+1):
#             c=ws.cell(1,ci2); c.fill=fl(bg); c.border=bd()
#     ws.row_dimensions[1].height = 22

#     # Row 2: column name headers
#     for ci, col in enumerate(final_cols, 1):
#         grp = col_group.get(col, 'info')
#         bg, _ = GROUP_STYLE[grp]
#         c = ws.cell(2, ci, col)
#         c.font=hf(9); c.fill=fl(bg); c.alignment=ctr(); c.border=bd()
#     ws.row_dimensions[2].height = 44

#     # Data rows
#     data = wide[final_cols].values.tolist()
#     for ri, row_vals in enumerate(data):
#         er = ri + 3
#         alt = ri % 2 == 0
#         for ci, (col, val) in enumerate(zip(final_cols, row_vals), 1):
#             grp = col_group.get(col, 'info')
#             _, alt_bg = GROUP_STYLE[grp]
#             # Convert numpy int64 to plain int for openpyxl
#             if hasattr(val, 'item'):
#                 val = val.item()
#             c = ws.cell(er, ci, val)
#             c.border = bd()
#             if grp == 'info':
#                 c.font = df_(9, bold=(col in ('RTO Code','RTO Name')))
#                 c.alignment = lft()
#                 if alt: c.fill = fl(alt_bg)
#             elif col == 'TOTAL REGISTRATIONS':
#                 c.font = df_(9, bold=True); c.fill = fl(alt_bg)
#                 c.alignment = ctr(); c.number_format = '#,##0'
#             else:
#                 c.font = df_(9); c.alignment = ctr(); c.number_format = '#,##0'
#                 if alt: c.fill = fl(alt_bg)
#         ws.row_dimensions[er].height = 14

#     ws.freeze_panes = ws.cell(3, len(ID_COLS) + 1)
#     auto_w(ws)
#     for col_letter, w in zip('ABCDEFG', [10, 7, 7, 12, 12, 10, 34]):
#         ws.column_dimensions[col_letter].width = w
#     ws.sheet_properties.tabColor = '1F3864'
#     wb.save(output_path)
#     log.info(f"Saved: {output_path}  ({round(output_path.stat().st_size/1024)} KB)")

# # ── Public API ────────────────────────────────────────────────────────────────
# def format_data(input_path, output_path, state_filter=None):
#     log.info(f"Loading: {input_path}")
#     df = pd.read_excel(input_path)
#     df = df.copy()
#     df['registrations_count'] = pd.to_numeric(df['registrations_count'], errors='coerce').fillna(0).astype(int)

#     if state_filter:
#         sf = [s.upper() for s in state_filter]
#         df = df[df['state_code'].str.upper().isin(sf)].copy()
#         log.info(f"Filtered to: {sf} -> {len(df):,} rows")

#     if df.empty:
#         log.error("No data after filtering.")
#         return None

#     df['rto_clean']   = df['rto_name'].apply(clean_rto)
#     df['state_clean'] = df['state_name'].apply(clean_state)

#     log.info(f"States : {', '.join(sorted(df['state_clean'].unique()))}")
#     log.info(f"RTOs   : {df['rto_clean'].nunique()}  |  Rows: {len(df):,}")
#     build_wide_sheet(df, Path(output_path))
#     return output_path

# # ── CLI ───────────────────────────────────────────────────────────────────────
# def main():
#     p = argparse.ArgumentParser(description='VAHAN Data Formatter')
#     p.add_argument('--input',  default=None)
#     p.add_argument('--output', default=None)
#     p.add_argument('--states', nargs='*', default=None)
#     args = p.parse_args()

#     data_dir = Path('vahan_data')
#     input_path = Path(args.input) if args.input else data_dir / 'vahan_registrations.xlsx'
#     if not input_path.exists():
#         log.error(f"Not found: {input_path}"); sys.exit(1)

#     if args.output:
#         output_path = Path(args.output)
#     else:
#         if args.states:
#             codes = '_'.join(sorted(s.upper() for s in args.states))
#         else:
#             try:
#                 codes = '_'.join(sorted(
#                     pd.read_excel(input_path, usecols=['state_code'])['state_code'].unique()))
#             except Exception:
#                 codes = 'ALL'
#         output_path = data_dir / f"VAHAN_{codes}_{datetime.now().strftime('%Y%m%d')}_formatted.xlsx"

#     data_dir.mkdir(exist_ok=True)
#     format_data(input_path, output_path, args.states)

# if __name__ == '__main__':
#     main()

# -*- coding: utf-8 -*-
"""
VAHAN Data Formatter — single wide sheet output.
Each row = RTO × Month. Columns = ID + all fuel types + vehicle categories +
vehicle classes + emission norms + makers + TOTAL.

USAGE:
    python vahan_format.py
    python vahan_format.py --input data.xlsx --output report.xlsx
    python vahan_format.py --states DL UP
"""
import re, sys, argparse, logging
from pathlib import Path
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)])
log = logging.getLogger(__name__)

MONTH_ORDER = ['JAN','FEB','MAR','APR','MAY','JUN','JUL','AUG','SEP','OCT','NOV','DEC']
ID_COLS = ['Period','Month','Year','State Code','State','RTO Code','RTO Name']

BREAKDOWNS = [
    ('fuel',             'fuel',             'BY FUEL TYPE'),
    ('vehicle_category', 'vehicle_category', 'BY VEHICLE CATEGORY'),
    ('vehicle_class',    'vehicle_class',    'BY VEHICLE CLASS'),
    ('norms',            'norms',            'BY EMISSION NORMS'),
    ('maker',            'maker',            'BY MAKER / BRAND'),
]
GROUP_LABELS = {
    'info':'IDENTIFICATION','fuel':'BY FUEL TYPE','vehicle_category':'BY VEHICLE CATEGORY',
    'vehicle_class':'BY VEHICLE CLASS','norms':'BY EMISSION NORMS',
    'maker':'BY MAKER / BRAND','total':'TOTAL',
}
GROUP_STYLE = {
    'info':             ('1F3864','D6E4F7'),
    'fuel':             ('833C00','FCE4D6'),
    'vehicle_category': ('375623','E2EFDA'),
    'vehicle_class':    ('4B2F8A','EAE0F7'),
    'norms':            ('1F4E79','DDEBF7'),
    'maker':            ('5C3317','F2E0D0'),
    'total':            ('404040','C6EFCE'),
}

# ── Style helpers ─────────────────────────────────────────────────────────────
def hf(sz=10):              return Font(bold=True,color='FFFFFF',name='Calibri',size=sz)
def df_(sz=9,bold=False):   return Font(name='Calibri',size=sz,bold=bold)
def fl(c):                  return PatternFill('solid',fgColor=c)
def bd():
    s=Side(style='thin',color='C0C0C0')
    return Border(left=s,right=s,top=s,bottom=s)
def ctr(): return Alignment(horizontal='center',vertical='center',wrap_text=False)
def lft(): return Alignment(horizontal='left',  vertical='center',wrap_text=False)

def auto_w(ws,min_w=6,max_w=28):
    for col in ws.columns:
        w=max(min_w,min(max_w,max((len(str(c.value)) if c.value is not None else 0) for c in col)+2))
        ws.column_dimensions[get_column_letter(col[0].column)].width=w

def clean_rto(name):
    m=re.match(r'^(.+?)\s*-\s*([A-Z]+\d+)\s*\(.*\)$',str(name).strip())
    return f"{m.group(2)} \u2013 {m.group(1).strip().title()}" if m else str(name).strip()

def clean_state(name):
    m=re.match(r'^(.+?)\s*\(',str(name).strip())
    return m.group(1).strip() if m else str(name).strip()

# ── Core builder ──────────────────────────────────────────────────────────────
def build_wide_sheet(df: pd.DataFrame, output_path: Path):
    df = df.copy()

    # Base grid: one row per RTO × month
    base = (df[['time_period','month','year','state_code','state_clean','rto_code','rto_clean']]
            .drop_duplicates()
            .copy()
            .rename(columns={'time_period':'Period','month':'Month','year':'Year',
                             'state_code':'State Code','state_clean':'State',
                             'rto_code':'RTO Code','rto_clean':'RTO Name'}))
    base = base.assign(_mn=base['Month'].map({m:i for i,m in enumerate(MONTH_ORDER)}))
    base = base.sort_values(['RTO Code','Year','_mn']).drop(columns='_mn').reset_index(drop=True)

    # Pivot each breakdown and merge onto base
    group_cols: dict[str, list] = {}
    wide = base.copy()

    for bt, cat_col, _ in BREAKDOWNS:
        sub = df[df['breakdown_type'] == bt].copy()
        if sub.empty:
            continue
        # Drop rows where the category value is purely numeric (scraper artefact —
        # these are RTO codes that leaked into the category column).
        # Note: vehicle_class entries start with "-" (e.g. "-Motor Car") so we
        # only drop values that are entirely digits, not ones starting with "-".
        sub = sub[~sub[cat_col].astype(str).str.strip()
                  .str.fullmatch(r'[0-9]+').fillna(False)].copy()
        if sub.empty:
            continue
        piv = (sub.pivot_table(
                    index=['time_period','rto_code'],
                    columns=cat_col,
                    values='registrations_count',
                    aggfunc='sum',
                    fill_value=0)
               .reset_index())
        piv.columns.name = None
        piv = piv.rename(columns={'time_period':'Period','rto_code':'RTO Code'})
        cat_names = [c for c in piv.columns if c not in ('Period','RTO Code')]
        group_cols[bt] = cat_names
        wide = wide.merge(piv, on=['Period','RTO Code'], how='left')

    # Fill NaN → 0 for all category columns
    all_cat_cols = [c for cols in group_cols.values() for c in cols]
    # Fix all numeric columns in one shot using assign — avoids pandas CoW warnings
    num_updates = {c: pd.to_numeric(wide[c], errors='coerce').fillna(0).astype(int)
                   for c in all_cat_cols if c in wide.columns}
    wide = wide.assign(**num_updates)

    # TOTAL = sum of fuel cols (or vehicle_category as fallback)
    total_src = [c for c in (group_cols.get('fuel') or group_cols.get('vehicle_category') or [])
                 if c in wide.columns]
    wide = wide.assign(**{'TOTAL REGISTRATIONS': wide[total_src].sum(axis=1) if total_src else 0})
    group_cols['total'] = ['TOTAL REGISTRATIONS']

    # Build final column list — only include cols that actually exist in wide
    final_cols = [c for c in ID_COLS if c in wide.columns]
    for bt, _, _ in BREAKDOWNS:
        final_cols += [c for c in group_cols.get(bt, []) if c in wide.columns]
    final_cols += ['TOTAL REGISTRATIONS']

    # Map each column to its group
    col_group = {c: 'info' for c in ID_COLS}
    for bt, _, _ in BREAKDOWNS:
        for c in group_cols.get(bt, []):
            col_group[c] = bt
    col_group['TOTAL REGISTRATIONS'] = 'total'

    # Safety: use reindex to guarantee all final_cols exist (fill missing with 0)
    wide = wide.reindex(columns=final_cols, fill_value=0)

    log.info(f"  Wide table: {len(wide)} rows x {len(final_cols)} columns")
    for bt, _, label in BREAKDOWNS:
        n = len([c for c in group_cols.get(bt, []) if c in wide.columns])
        if n: log.info(f"    {label:28}: {n} cols")

    # ── Write Excel ────────────────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = 'Vehicle Registration Data'

    # Find column index ranges per group
    grp_ranges: dict[str, tuple] = {}
    for ci, col in enumerate(final_cols, 1):
        grp = col_group.get(col, 'info')
        grp_ranges[grp] = (grp_ranges[grp][0], ci) if grp in grp_ranges else (ci, ci)

    # Row 1: merged group header cells
    for grp, (s, e) in grp_ranges.items():
        bg, _ = GROUP_STYLE[grp]
        sl, el = get_column_letter(s), get_column_letter(e)
        if s != e:
            ws.merge_cells(f'{sl}1:{el}1')
        cell = ws.cell(1, s, GROUP_LABELS.get(grp, grp.upper()))
        cell.font=hf(10); cell.fill=fl(bg); cell.alignment=ctr(); cell.border=bd()
        for ci2 in range(s+1, e+1):
            c=ws.cell(1,ci2); c.fill=fl(bg); c.border=bd()
    ws.row_dimensions[1].height = 22

    # Row 2: column name headers
    for ci, col in enumerate(final_cols, 1):
        grp = col_group.get(col, 'info')
        bg, _ = GROUP_STYLE[grp]
        c = ws.cell(2, ci, col)
        c.font=hf(9); c.fill=fl(bg); c.alignment=ctr(); c.border=bd()
    ws.row_dimensions[2].height = 44

    # Data rows
    data = wide[final_cols].values.tolist()
    for ri, row_vals in enumerate(data):
        er = ri + 3
        alt = ri % 2 == 0
        for ci, (col, val) in enumerate(zip(final_cols, row_vals), 1):
            grp = col_group.get(col, 'info')
            _, alt_bg = GROUP_STYLE[grp]
            # Convert numpy int64 to plain int for openpyxl
            if hasattr(val, 'item'):
                val = val.item()
            c = ws.cell(er, ci, val)
            c.border = bd()
            if grp == 'info':
                c.font = df_(9, bold=(col in ('RTO Code','RTO Name')))
                c.alignment = lft()
                if alt: c.fill = fl(alt_bg)
            elif col == 'TOTAL REGISTRATIONS':
                c.font = df_(9, bold=True); c.fill = fl(alt_bg)
                c.alignment = ctr(); c.number_format = '#,##0'
            else:
                c.font = df_(9); c.alignment = ctr(); c.number_format = '#,##0'
                if alt: c.fill = fl(alt_bg)
        ws.row_dimensions[er].height = 14

    ws.freeze_panes = ws.cell(3, len(ID_COLS) + 1)
    auto_w(ws)
    for col_letter, w in zip('ABCDEFG', [10, 7, 7, 12, 12, 10, 34]):
        ws.column_dimensions[col_letter].width = w
    ws.sheet_properties.tabColor = '1F3864'
    wb.save(output_path)
    log.info(f"Saved: {output_path}  ({round(output_path.stat().st_size/1024)} KB)")

# ── Public API ────────────────────────────────────────────────────────────────
def format_data(input_path, output_path, state_filter=None):
    log.info(f"Loading: {input_path}")
    df = pd.read_excel(input_path)
    df = df.copy()
    df['registrations_count'] = pd.to_numeric(df['registrations_count'], errors='coerce').fillna(0).astype(int)

    if state_filter:
        sf = [s.upper() for s in state_filter]
        df = df[df['state_code'].str.upper().isin(sf)].copy()
        log.info(f"Filtered to: {sf} -> {len(df):,} rows")

    if df.empty:
        log.error("No data after filtering.")
        return None

    df['rto_clean']   = df['rto_name'].apply(clean_rto)
    df['state_clean'] = df['state_name'].apply(clean_state)

    log.info(f"States : {', '.join(sorted(df['state_clean'].unique()))}")
    log.info(f"RTOs   : {df['rto_clean'].nunique()}  |  Rows: {len(df):,}")
    build_wide_sheet(df, Path(output_path))
    return output_path

# ── CLI ───────────────────────────────────────────────────────────────────────
def main():
    p = argparse.ArgumentParser(description='VAHAN Data Formatter')
    p.add_argument('--input',  default=None)
    p.add_argument('--output', default=None)
    p.add_argument('--states', nargs='*', default=None)
    args = p.parse_args()

    data_dir = Path('vahan_data')
    input_path = Path(args.input) if args.input else data_dir / 'vahan_registrations.xlsx'
    if not input_path.exists():
        log.error(f"Not found: {input_path}"); sys.exit(1)

    if args.output:
        output_path = Path(args.output)
    else:
        if args.states:
            codes = '_'.join(sorted(s.upper() for s in args.states))
        else:
            try:
                codes = '_'.join(sorted(
                    pd.read_excel(input_path, usecols=['state_code'])['state_code'].unique()))
            except Exception:
                codes = 'ALL'
        output_path = data_dir / f"VAHAN_{codes}_{datetime.now().strftime('%Y%m%d')}_formatted.xlsx"

    data_dir.mkdir(exist_ok=True)
    format_data(input_path, output_path, args.states)

if __name__ == '__main__':
    main()
