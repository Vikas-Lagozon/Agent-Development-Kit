# -*- coding: utf-8 -*-
"""
VAHAN Dashboard — Combined Pipeline
=====================================
Step 1: Scrape VAHAN dashboard (Selenium)
Step 2: Format into wide Excel  
Step 3: Export BigQuery side-by-side format

USAGE:
    python vahan_pipeline.py --headed --states LA
    python vahan_pipeline.py --headed --states DL CH LA UP
    python vahan_pipeline.py                              # ALL states
    python vahan_pipeline.py --reset --headed --states LA # clear & redo
    python vahan_pipeline.py --format-only                # skip scrape
"""

import time, json, sys, argparse, logging, re
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options as ChromeOptions
from selenium.webdriver.firefox.options import Options as FirefoxOptions
from selenium.common.exceptions import (
    TimeoutException, NoSuchElementException, StaleElementReferenceException
)

# ── Config ───────────────────────────────────────────────────────────────────
BASE_URL   = "https://vahan.parivahan.gov.in/vahan4dashboard/"
OUT_DIR    = Path("vahan_data")

START_YEAR, START_MONTH = 2025, 1
NOW = datetime.now()
END_YEAR, END_MONTH = NOW.year, NOW.month

WAIT_PAGE   = 18
WAIT_STATE  = 8
WAIT_RTO    = 14
WAIT_YEAR   = 12
WAIT_MONTH  = 12   # time for the month breakdown panel to fully render
RETRIES     = 3

MONTH_ABBRS = ["JAN","FEB","MAR","APR","MAY","JUN",
               "JUL","AUG","SEP","OCT","NOV","DEC"]

SECTION_KEYWORDS = {
    "vehicle_class":    "vehicle class",
    "vehicle_category": "vehicle category",
    "fuel":             "fuel",
    "norms":            "norms",
    "maker":            "maker",
}

# ── Logging ──────────────────────────────────────────────────────────────────
OUT_DIR.mkdir(exist_ok=True)

class AsciiFilter(logging.Filter):
    def filter(self, record):
        record.msg = str(record.msg).encode("ascii","replace").decode("ascii")
        return True

handlers = [
    logging.FileHandler(OUT_DIR/"vahan_scraper.log", encoding="utf-8"),
    logging.StreamHandler(sys.stdout),
]
for h in handlers:
    h.addFilter(AsciiFilter())
logging.basicConfig(level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s", handlers=handlers)
log = logging.getLogger(__name__)

PROGRESS_FILE = OUT_DIR / "vahan_progress.json"
RAW_FILE      = OUT_DIR / "vahan_raw.json"

def load_json(p, default):
    return json.loads(p.read_text(encoding="utf-8")) if p.exists() else default

def save_json(p, obj):
    p.write_text(json.dumps(obj, indent=2, ensure_ascii=False), encoding="utf-8")

def append_raw(rec):
    d = load_json(RAW_FILE, [])
    d.append(rec)
    save_json(RAW_FILE, d)

# ── Browser ──────────────────────────────────────────────────────────────────
def get_driver(browser="chrome", headed=False):
    if browser == "chrome":
        opts = ChromeOptions()
        if not headed:
            opts.add_argument("--headless=new")
        for a in ["--no-sandbox","--disable-dev-shm-usage","--disable-gpu",
                  "--window-size=1920,1200","--disable-blink-features=AutomationControlled"]:
            opts.add_argument(a)
        opts.add_experimental_option("excludeSwitches", ["enable-automation"])
        opts.add_experimental_option("useAutomationExtension", False)
        opts.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36")
        d = webdriver.Chrome(options=opts)
    else:
        opts = FirefoxOptions()
        if not headed: opts.add_argument("--headless")
        d = webdriver.Firefox(options=opts)
    d.set_page_load_timeout(60)
    return d

# ── JS helpers ───────────────────────────────────────────────────────────────
def js(d, s, *a):
    try: return d.execute_script(s, *a)
    except Exception as e: log.debug(f"js: {e}"); return None

def wait_jquery(d, timeout=20):
    try:
        WebDriverWait(d, timeout).until(
            lambda x: x.execute_script(
                "return typeof jQuery!=='undefined' && jQuery.active===0;"))
    except Exception: pass

# ── Dynamic ID discovery ─────────────────────────────────────────────────────
# JSF regenerates element IDs (j_idt30, j_idt43, j_idt48) on every app
# restart. We discover them at runtime by inspecting option values/counts.

_ID_CACHE: dict = {}  # cached once per session

def _discover_ids(d):
    """
    Scan all <select> and <button> elements to identify:
      - type_id:   select whose options include 'Y' (yearwise) and 'M' (monthly)
      - state_id:  select with 30+ options (state list)
      - rto_id:    select named 'selectedRto' or with id containing 'Rto'
      - refresh_id: button/input that triggers data refresh
    Falls back to known IDs if discovery fails.
    """
    if _ID_CACHE:
        return _ID_CACHE

    result = js(d, """
        var selects = document.querySelectorAll('select');
        var found = {type_id:null, state_id:null, rto_id:null, refresh_id:null};
        for (var i=0; i<selects.length; i++) {
            var s = selects[i];
            var id = s.id || '';
            var name = (s.name || '').toLowerCase();
            var opts = s.options;
            var vals = [];
            for (var j=0; j<opts.length; j++) vals.push(opts[j].value);

            // RTO select: id/name contains 'rto' (case-insensitive)
            if (id.toLowerCase().indexOf('rto') !== -1 || name.indexOf('rto') !== -1) {
                found.rto_id = id; continue;
            }
            // Type select: has options Y, M, or A (small option set)
            if (vals.indexOf('Y') !== -1 || vals.indexOf('A') !== -1) {
                if (opts.length < 10) { found.type_id = id; continue; }
            }
            // State select: 30-40 options with short 2-letter values
            if (opts.length >= 25 && opts.length <= 45) {
                var short = vals.filter(function(v){ return v && v.length<=3 && v!=='-1'; });
                if (short.length >= 20) { found.state_id = id; continue; }
            }
        }
        // Refresh button: find button nearest to the RTO/state selects
        // It's typically a button right after the dropdowns in the form
        var allBtns = document.querySelectorAll('button,input[type=submit],input[type=button]');
        var candidates = [];
        for (var k=0; k<allBtns.length; k++) {
            var bid = allBtns[k].id || '';
            var btext = (allBtns[k].innerText || allBtns[k].value || '').trim().toLowerCase();
            var bcls = (allBtns[k].className || '').toLowerCase();
            if (!bid) continue;
            // Skip obvious non-refresh buttons
            if (btext.indexOf('cancel') !== -1 || btext.indexOf('close') !== -1 ||
                btext.indexOf('reset') !== -1 || bcls.indexOf('ui-datepicker') !== -1) continue;
            // Good candidates: small/icon buttons with idt in id, or explicit refresh/search
            if (bid.indexOf('idt') !== -1) {
                candidates.push({id: bid, text: btext, len: btext.length});
            }
        }
        // Prefer buttons with refresh/search text, then shortest text (icon buttons)
        candidates.sort(function(a,b) {
            var aGood = a.text.indexOf('refresh')!==-1 || a.text.indexOf('search')!==-1 ? 0 : 1;
            var bGood = b.text.indexOf('refresh')!==-1 || b.text.indexOf('search')!==-1 ? 0 : 1;
            if (aGood !== bGood) return aGood - bGood;
            return a.len - b.len;
        });
        if (candidates.length > 0) found.refresh_id = candidates[0].id;
        return found;
    """)

    # Fallbacks to known IDs
    ids = {
        'type_id':    (result or {}).get('type_id')    or 'j_idt30_input',
        'state_id':   (result or {}).get('state_id')   or 'j_idt43_input',
        'rto_id':     (result or {}).get('rto_id')     or 'selectedRto_input',
        'refresh_id': (result or {}).get('refresh_id') or 'j_idt48',
    }
    log.info(f"  Element IDs: {ids}")
    _ID_CACHE.update(ids)
    return _ID_CACHE

def _get_id(d, key):
    return _discover_ids(d)[key]

def pf_set(d, eid, val, wait=4):
    # If eid is a logical key, resolve it; otherwise use directly
    resolved = _ID_CACHE.get(eid, eid)
    r = js(d, """
        var s=document.getElementById(arguments[0]);
        if(!s) return 'NOT_FOUND';
        s.value=arguments[1];
        s.dispatchEvent(new Event('change',{bubbles:true}));
        return s.value;
    """, resolved, val)
    time.sleep(wait); wait_jquery(d); return r

def get_opts(d, eid):
    resolved = _ID_CACHE.get(eid, eid)
    return js(d, """
        var s=document.getElementById(arguments[0]);
        if(!s) return [];
        var r=[];
        for(var i=0;i<s.options.length;i++){
            var v=s.options[i].value,t=s.options[i].text.trim();
            if(v&&v!=='-1') r.push([v,t]);
        }
        return r;
    """, resolved) or []

def click_refresh(d, wait=WAIT_RTO):
    """
    Click the refresh/search button. Tries multiple strategies:
    1. Known/discovered ID via WebDriverWait
    2. JS click on discovered ID
    3. Scan ALL buttons and click the one that triggers data reload
    """
    rid = _get_id(d, 'refresh_id')

    # Strategy 1: Selenium click on discovered ID
    try:
        b = WebDriverWait(d, 8).until(EC.element_to_be_clickable((By.ID, rid)))
        js(d, "arguments[0].click();", b)
        time.sleep(wait); wait_jquery(d)
        return
    except Exception as e:
        log.debug(f"  Refresh strategy 1 failed (id={rid}): {e}")

    # Strategy 2: JS click on discovered ID (bypasses interactability check)
    try:
        result = js(d, """
            var b = document.getElementById(arguments[0]);
            if (b) { b.click(); return 'ok'; }
            return 'not_found';
        """, rid)
        if result == 'ok':
            time.sleep(wait); wait_jquery(d)
            return
    except Exception as e:
        log.debug(f"  Refresh strategy 2 failed: {e}")

    # Strategy 3: Find any button/input that looks like a refresh/search trigger
    # by clicking all visible buttons with idt in their id and checking which one
    # causes the year links panel to update
    log.info("  Refresh: trying button scan strategy...")
    clicked = js(d, """
        var btns = document.querySelectorAll('button, input[type=submit], input[type=button], a.ui-button');
        for (var i=0; i<btns.length; i++) {
            var bid = btns[i].id || '';
            var bcls = btns[i].className || '';
            var btxt = (btns[i].innerText || btns[i].value || '').trim().toLowerCase();
            // Skip dropdowns, cancel, close buttons
            if (btxt.indexOf('cancel') !== -1 || btxt.indexOf('close') !== -1) continue;
            // Click buttons with idt in id that have no text (icon buttons) or say refresh/search/show
            if (bid && bid.indexOf('idt') !== -1) {
                if (btxt === '' || btxt.indexOf('refresh') !== -1 || 
                    btxt.indexOf('search') !== -1 || btxt.indexOf('show') !== -1 ||
                    btxt.indexOf('go') !== -1 || btxt.length < 4) {
                    btns[i].click();
                    return 'clicked:' + bid + ':' + btxt;
                }
            }
        }
        return null;
    """)
    if clicked:
        log.info(f"  Refresh strategy 3: {clicked}")
        # Update the cache with the real button ID
        real_id = clicked.split(':')[1] if ':' in clicked else rid
        _ID_CACHE['refresh_id'] = real_id
        time.sleep(wait); wait_jquery(d)
        return

    log.warning(f"  All refresh strategies failed — page may show stale data")

# ── Text helpers ──────────────────────────────────────────────────────────────
def stext(el):
    try: return el.text.strip()
    except StaleElementReferenceException: return ""

def parse_int(s):
    try: return int(str(s).replace(",","").replace(" ","").strip())
    except (ValueError, AttributeError): return None

# ── TARGETED panel discovery ──────────────────────────────────────────────────
#
# After clicking a month tab, PrimeFaces AJAX updates a panel.
# The updated panel has a header/title containing the month name.
# We find that specific panel, then extract ONLY from tables inside it.
#
# The panel title typically looks like:
#   "Total Registration Data of DWARKA - DL9, Delhi (2025 Jan)"
#   or "Registration Data JAN 2025" etc.
#
# We identify it by looking for an element whose text contains the month abbr.
# ─────────────────────────────────────────────────────────────────────────────

def find_active_detail_panel(driver, month_abbr: str, year: int):
    """
    Find the DOM container that was updated after clicking the month tab.
    Returns the element, or None.
    Priority:
    1. Element with text containing month_abbr + year (most specific)
    2. The infoMsg / panelHeader panels if they contain the month
    3. Any panel/div whose text starts with month_abbr
    """
    abbr_u = month_abbr.upper()
    yr_str  = str(year)

    # Try known panel IDs first
    for panel_id in ["infoMsg", "panelHeader", "yearWiseRegnDataTable",
                     "mainpagepnl", "dashboardContentsPanel"]:
        try:
            el = driver.find_element(By.ID, panel_id)
            txt = (el.get_attribute("innerText") or "").upper()
            if abbr_u in txt and yr_str in txt:
                log.info(f"        Active panel found: #{panel_id}")
                return el
        except NoSuchElementException:
            continue

    # Generic: walk from pnl_regn_content siblings/parent
    try:
        base = driver.find_element(By.ID, "pnl_regn_content")
        parent = base.find_element(By.XPATH, "..")
        siblings = parent.find_elements(By.XPATH, "./*")
        for sib in siblings:
            txt = (sib.get_attribute("innerText") or "").upper()
            if abbr_u in txt and yr_str in txt and len(sib.find_elements(By.TAG_NAME,"table")) > 0:
                log.info(f"        Active panel found: sibling of pnl_regn_content")
                return sib
    except Exception:
        pass

    # Broader: any div/section whose innerText contains month+year and has tables
    panels = driver.find_elements(By.XPATH,
        "//div[contains(@class,'ui-panel') or contains(@class,'ui-widget')]")
    for p in panels:
        try:
            txt = (p.get_attribute("innerText") or "").upper()
            if abbr_u in txt and yr_str in txt:
                tables = p.find_elements(By.TAG_NAME,"table")
                if tables:
                    log.info(f"        Active panel found: ui-panel/widget with {len(tables)} tables")
                    return p
        except Exception:
            continue

    log.warning(f"        No panel found containing '{abbr_u}' + '{yr_str}' — will scan full page")
    return None

def table_header_text(tbl, driver=None) -> str:
    """Get the header text of a table. Uses JS innerText for accuracy."""
    try:
        # Use JS to get just the first row's text cleanly
        if driver:
            hdr = driver.execute_script("""
                var tbl = arguments[0];
                var rows = tbl.rows;
                if (!rows || rows.length === 0) return '';
                // Check thead first
                var thead = tbl.querySelector('thead');
                if (thead) {
                    return (thead.innerText || '').trim().toLowerCase();
                }
                // Fallback: first row
                return (rows[0].innerText || '').trim().toLowerCase();
            """, tbl)
            return hdr or ""
        ths = tbl.find_elements(By.CSS_SELECTOR,"thead th,thead td")
        if ths: return " ".join(stext(t) for t in ths).lower()
        trs = tbl.find_elements(By.TAG_NAME,"tr")
        if trs:
            cells = trs[0].find_elements(By.TAG_NAME,"td") + \
                    trs[0].find_elements(By.TAG_NAME,"th")
            return " ".join(stext(c) for c in cells).lower()
    except Exception: pass
    return ""

def _classify_table(tbl, driver) -> str:
    """
    Return the section key for a table by checking:
    1. Table header row (most specific - e.g. "Vehicle Class(2026)")
    2. Nearest preceding heading/caption element
    3. Immediate parent's own text (excluding child table text)
    Returns one of: vehicle_class, vehicle_category, fuel, norms, maker, or ''
    """
    # SECTION_KEYWORDS ordered so longer/more specific come first
    # "vehicle class" must be checked BEFORE "vehicle" to avoid false matches
    ordered = [
        ("vehicle_class",    "vehicle class"),
        ("vehicle_category", "vehicle category"),
        ("fuel",             "fuel"),
        ("norms",            "norms"),
        ("maker",            "maker"),
    ]

    # 1. Check table header via JS (most reliable)
    hdr = table_header_text(tbl, driver)
    for key, kw in ordered:
        if kw in hdr:
            return key

    # 2. Check nearest preceding sibling heading or caption
    try:
        label = driver.execute_script("""
            var tbl = arguments[0];
            // Check caption inside table
            var cap = tbl.querySelector('caption');
            if (cap) return (cap.innerText||'').toLowerCase();
            // Walk up and look for preceding text node / heading
            var el = tbl;
            for (var i=0; i<5; i++) {
                el = el.parentElement;
                if (!el) break;
                // Look for heading elements before this table
                var kids = el.children;
                for (var j=0; j<kids.length; j++) {
                    var tag = kids[j].tagName.toLowerCase();
                    if (kids[j].contains(tbl)) break;
                    if (['h1','h2','h3','h4','h5','h6','span','div','p','th','td'].includes(tag)) {
                        var txt = (kids[j].innerText||'').trim().toLowerCase();
                        if (txt && txt.length < 60) return txt;
                    }
                }
                // Check parent's own direct text (not children)
                var directTxt = '';
                for (var n=0; n<el.childNodes.length; n++) {
                    if (el.childNodes[n].nodeType === 3) {
                        directTxt += el.childNodes[n].textContent;
                    }
                }
                directTxt = directTxt.trim().toLowerCase();
                if (directTxt && directTxt.length < 60) return directTxt;
            }
            return '';
        """, tbl)
        if label:
            for key, kw in ordered:
                if kw in label:
                    return key
    except Exception:
        pass

    return ""

# Keywords that identify a row as a section sub-header (not real data)
_HEADER_KEYWORDS = {
    "vehicle class", "vehicle category", "fuel type", "fuel",
    "norms", "maker brand", "maker/brand", "total", "sr.no",
}

def _is_header_row(tr) -> bool:
    """
    Return True if this <tr> is a table header row that should be skipped.
    A row is a header if:
      1. It lives inside a <thead> element, OR
      2. All its cells are <th> (not <td>), OR
      3. Its first cell text exactly matches a known section keyword
    """
    try:
        # Check if inside <thead>
        parent_tag = tr.find_element(By.XPATH, "..").tag_name.lower()
        if parent_tag == "thead":
            return True
        # Check if all cells are <th>
        cells = tr.find_elements(By.TAG_NAME, "th")
        tds   = tr.find_elements(By.TAG_NAME, "td")
        if cells and not tds:
            return True
        # Check if first cell text is a section keyword
        first_cells = cells or tds
        if first_cells:
            txt = stext(first_cells[0]).lower().strip()
            if txt in _HEADER_KEYWORDS:
                return True
    except Exception:
        pass
    return False

def extract_table_rows(tbl) -> list[tuple[str,int]]:
    rows = []
    try:
        trs = tbl.find_elements(By.TAG_NAME,"tr")
        for tr in trs:
            # Skip genuine header rows — do NOT blindly skip row 0
            if _is_header_row(tr):
                continue
            tds = tr.find_elements(By.TAG_NAME,"td")
            if len(tds) < 2:
                continue
            name = stext(tds[0]).strip()
            if not name:
                continue
            # Skip section sub-header rows by keyword match
            if name.lower().strip() in _HEADER_KEYWORDS:
                continue
            # Extract the rightmost numeric value as the count
            count = None
            for td in reversed(tds):
                v = parse_int(stext(td))
                if v is not None:
                    count = v
                    break
            if name and count is not None and count >= 0:
                rows.append((name, count))
    except Exception as e:
        log.debug(f"extract_rows: {e}")
    return rows

def find_breakdown_in_scope(scope_el, driver=None) -> dict[str,list]:
    """Extract all 5 breakdown tables from a given container element."""
    result = {k:[] for k in SECTION_KEYWORDS}
    tables = scope_el.find_elements(By.TAG_NAME,"table")

    for tbl in tables:
        key = _classify_table(tbl, driver) if driver else ""
        if not key:
            # Fallback: old method using combined header + parent text
            hdr = table_header_text(tbl)
            try:
                parent_txt = (tbl.find_element(By.XPATH,"..").get_attribute("innerText") or "")[:200].lower()
            except Exception:
                parent_txt = ""
            combined = hdr + " " + parent_txt
            for k, kw in SECTION_KEYWORDS.items():
                if kw in combined:
                    key = k; break

        if key:
            rows = extract_table_rows(tbl)
            if len(rows) > len(result[key]):
                result[key] = rows

    # Heading-based fallback for any still-missing sections
    for key, kw in SECTION_KEYWORDS.items():
        if result[key]: continue
        try:
            headings = scope_el.find_elements(By.XPATH,
                f".//*[contains(translate(normalize-space(text()),'abcdefghijklmnopqrstuvwxyz',"
                f"'ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'{kw.upper()}')]")
            for h in headings:
                parent = h
                for _ in range(6):
                    try:
                        parent = parent.find_element(By.XPATH,"..")
                        tbls = parent.find_elements(By.TAG_NAME,"table")
                        if tbls:
                            rows = extract_table_rows(tbls[0])
                            if len(rows) > len(result[key]):
                                result[key] = rows
                            break
                    except Exception: break
        except Exception: pass

    return result

def extract_all_breakdowns(driver, month_abbr: str, year: int) -> dict[str,list]:
    """
    Find the active monthly panel and extract breakdowns ONLY from it.
    Falls back to full-page scan only if no specific panel found.
    """
    panel = find_active_detail_panel(driver, month_abbr, year)

    if panel:
        result = find_breakdown_in_scope(panel, driver)
        for k, rows in result.items():
            log.info(f"        {k}: {len(rows)} rows")
        return result

    # FALLBACK: full page scan — but skip the pnl_regn_content "Till Today" tables
    # by excluding containers that contain "till today" in their text
    log.info("        [fallback] scanning full page, excluding Till Today panel...")
    result = {k:[] for k in SECTION_KEYWORDS}
    tables = driver.find_elements(By.TAG_NAME,"table")
    for tbl in tables:
        try:
            # Skip if this table is inside the "Till Today" persistent section
            ancestor_text = js(driver, """
                var el = arguments[0];
                for (var i=0; i<10; i++) {
                    el = el.parentElement;
                    if (!el) break;
                    if ((el.innerText||'').toLowerCase().indexOf('till today') !== -1 &&
                        (el.innerText||'').toLowerCase().indexOf('""" + month_abbr.lower() + """') === -1) {
                        return 'skip';
                    }
                }
                return 'ok';
            """, tbl)
            if ancestor_text == "skip":
                continue

            hdr = table_header_text(tbl)
            for key, kw in SECTION_KEYWORDS.items():
                if kw in hdr:
                    rows = extract_table_rows(tbl)
                    if len(rows) > len(result[key]):
                        result[key] = rows
                    break
        except StaleElementReferenceException:
            continue
        except Exception:
            continue

    for k, rows in result.items():
        log.info(f"        {k}: {len(rows)} rows")
    return result

# ── Year / Month navigation ────────────────────────────────────────────────────
def get_year_links(driver):
    try:
        panel = driver.find_element(By.ID,"pnl_regn_content")
        return [stext(l) for l in panel.find_elements(By.CSS_SELECTOR,"a.ui-commandlink") if stext(l)]
    except Exception: return []

def click_year(driver, year: int) -> bool:
    target = str(year)
    try:
        panel = driver.find_element(By.ID,"pnl_regn_content")
        for lnk in panel.find_elements(By.CSS_SELECTOR,"a.ui-commandlink"):
            t = stext(lnk)
            if target in t and "Till" not in t:
                js(driver,"arguments[0].click();",lnk)
                time.sleep(WAIT_YEAR); wait_jquery(driver); return True
    except Exception: pass

    found = js(driver,"""
        var links=document.querySelectorAll('#pnl_regn_content a.ui-commandlink,a.ui-commandlink');
        for(var i=0;i<links.length;i++){
            var t=(links[i].innerText||'').trim();
            if(t.indexOf(arguments[0])!==-1 && t.indexOf('Till')===-1 && t.length<10){
                links[i].click(); return 'ok:'+t;
            }
        }
        return null;
    """, target)
    if found:
        time.sleep(WAIT_YEAR); wait_jquery(driver); return True
    return False

def wait_for_month_tabs(driver, timeout=15) -> bool:
    """
    Wait up to `timeout` seconds for month tab elements to appear on the page.
    Returns True if found, False if timed out.
    """
    deadline = time.time() + timeout
    while time.time() < deadline:
        tabs = _scan_month_tabs(driver)
        if tabs:
            return True
        time.sleep(1)
    return False

def _scan_month_tabs(driver) -> list[tuple[str, object]]:
    """
    Find month tab <a> elements on the VAHAN dashboard.

    Diagnostic confirmed: after year click, month tabs are rendered as
      <a id="j_idt6XX" class="ui-commandlink ui-widget font-color">JAN</a>
    inside the 'panelHeader' / 'infoMsg' panel.

    Key issue: Selenium's el.text includes child text and may return
    "JAN\n4,227.00" instead of "JAN". We use JS innerText (trimmed first
    line only) to match exactly.
    """
    # Strategy 1: JS scan of all <a class="ui-commandlink"> — uses innerText
    # first line only so nested numbers don't pollute the match.
    results = js(driver, """
        var abbrs = {'JAN':1,'FEB':1,'MAR':1,'APR':1,'MAY':1,'JUN':1,
                     'JUL':1,'AUG':1,'SEP':1,'OCT':1,'NOV':1,'DEC':1};
        var els = document.querySelectorAll('a.ui-commandlink');
        var found = [];
        for (var i=0; i<els.length; i++) {
            var raw = (els[i].innerText || '').trim();
            var first = raw.split(/[\\n\\r\\t,]/)[0].trim();
            if (abbrs[first]) found.push([first, els[i].id]);
        }
        return found;
    """)

    if results:
        pairs = []
        for (abbr, eid) in results:
            try:
                el = driver.find_element(By.ID, eid)
                pairs.append((abbr, el))
            except Exception:
                pass
        if pairs:
            return pairs

    # Strategy 2: All <a> tags via JS innerText (catches any class variant)
    results2 = js(driver, """
        var abbrs = {'JAN':1,'FEB':1,'MAR':1,'APR':1,'MAY':1,'JUN':1,
                     'JUL':1,'AUG':1,'SEP':1,'OCT':1,'NOV':1,'DEC':1};
        var els = document.getElementsByTagName('a');
        var found = [];
        for (var i=0; i<els.length; i++) {
            var raw = (els[i].innerText || '').trim();
            var first = raw.split(/[\\n\\r\\t,]/)[0].trim();
            if (abbrs[first]) found.push([first, els[i].id]);
        }
        return found;
    """)

    if results2:
        pairs = []
        for (abbr, eid) in results2:
            try:
                el = driver.find_element(By.ID, eid) if eid else None
                if el:
                    pairs.append((abbr, el))
            except Exception:
                pass
        if pairs:
            return pairs

    # Strategy 3: Selenium fallback — search known panels for any clickable element
    ALL_TAGS = ("self::td or self::th or self::a or self::span or "
                "self::li or self::div or self::button or self::p")
    for sid in ["panelHeader", "infoMsg", "yearWiseRegnDataTable", "mainpagepnl"]:
        try:
            scope = driver.find_element(By.ID, sid)
            hits = []
            for el in scope.find_elements(By.XPATH, f".//*[{ALL_TAGS}]"):
                try:
                    # Use JS innerText to avoid child-text pollution
                    raw = driver.execute_script(
                        "return (arguments[0].innerText||'').trim().split(/[\\n\\r\\t,]/)[0].trim();", el)
                    if raw in MONTH_ABBRS:
                        hits.append((raw, el))
                except Exception:
                    pass
            if hits:
                return hits
        except NoSuchElementException:
            continue

    return []

def get_month_tabs(driver):
    """Find month tab elements after year click, with wait."""
    # First try immediately
    tabs = _scan_month_tabs(driver)
    if tabs:
        return tabs
    # Give it a bit more time and retry
    time.sleep(3)
    wait_jquery(driver)
    return _scan_month_tabs(driver)

def click_month_tab(driver, month_abbr: str) -> bool:
    tabs = get_month_tabs(driver)
    for (m, el) in tabs:
        if m == month_abbr:
            try:
                js(driver,"arguments[0].click();",el)
                time.sleep(WAIT_MONTH); wait_jquery(driver); return True
            except StaleElementReferenceException: pass

    # Pure JS click — uses innerText first-line to avoid child-text pollution
    found = js(driver,"""
        var abbr=arguments[0];
        var abbrs={'JAN':1,'FEB':1,'MAR':1,'APR':1,'MAY':1,'JUN':1,
                   'JUL':1,'AUG':1,'SEP':1,'OCT':1,'NOV':1,'DEC':1};
        var tags=['a','td','th','span','li','div','button','p','label'];
        for(var t=0;t<tags.length;t++){
            var els=document.getElementsByTagName(tags[t]);
            for(var i=0;i<els.length;i++){
                var raw=(els[i].innerText||els[i].value||'').trim();
                var first=raw.split(/[\\n\\r\\t,]/)[0].trim();
                if(first===abbr){
                    els[i].click(); return 'ok:'+tags[t]+':'+i;
                }
            }
        }
        return null;
    """, month_abbr)
    if found:
        log.info(f"        JS click: {found}")
        time.sleep(WAIT_MONTH); wait_jquery(driver); return True
    return False

# ── Scrape one RTO ────────────────────────────────────────────────────────────
def months_list():
    out, y, m = [], START_YEAR, START_MONTH
    while (y,m) <= (END_YEAR, END_MONTH):
        out.append((y, MONTH_ABBRS[m-1]))
        m += 1
        if m > 12: m, y = 1, y+1
    return out

def scrape_rto(driver, sc, sn, rc, rn, months) -> list[dict]:
    rows = []
    by_year: dict[int,list[str]] = {}
    for (yr, mo) in months: by_year.setdefault(yr,[]).append(mo)

    for year in sorted(by_year):
        avail = get_year_links(driver)
        log.info(f"      Year {year} | available links: {avail}")

        if not click_year(driver, year):
            log.warning(f"      Cannot click {year} — no data for this year")
            continue

        # Wait for month tabs to render (some states are slower / different UI)
        wait_for_month_tabs(driver, timeout=18)
        tabs = get_month_tabs(driver)
        log.info(f"      Month tabs ({len(tabs)} found): {[m for m,_ in tabs]}")
        if not tabs:
            hint = js(driver, """
                var c={};
                ['td','th','a','span','li','div','button','p'].forEach(function(t){
                    c[t]=document.getElementsByTagName(t).length;
                });
                return JSON.stringify(c);
            """)
            log.warning(f"      No month tabs — page element counts: {hint}")

        for month_abbr in by_year[year]:
            log.info(f"        Month {month_abbr} ...")
            tp = f"{year}-{MONTH_ABBRS.index(month_abbr)+1:02d}"
            base = dict(time_period=tp, month=month_abbr, year=str(year),
                        state_code=sc, state_name=sn, rto_code=rc, rto_name=rn)

            # Re-click year before every month tab — after long table extraction
            # (3-4 min) PrimeFaces session times out and collapses the month panel.
            click_year(driver, year)
            wait_for_month_tabs(driver, timeout=12)

            if not click_month_tab(driver, month_abbr):
                log.warning(f"        Cannot click {month_abbr} — retrying after year re-click")
                time.sleep(3)
                click_year(driver, year)
                wait_for_month_tabs(driver, timeout=15)
                if not click_month_tab(driver, month_abbr):
                    log.warning(f"        Cannot click {month_abbr} — skipping")
                    rows.append({**base,"breakdown_type":"total","maker":"",
                        "vehicle_class":"","vehicle_category":"","fuel":"","norms":"",
                        "registrations_count":None})
                    continue

            bkd = extract_all_breakdowns(driver, month_abbr, year)
            added = 0

            for section, items in bkd.items():
                for (name, count) in items:
                    row = {**base, "breakdown_type": section}
                    row["maker"]            = name if section=="maker"            else ""
                    row["vehicle_class"]    = name if section=="vehicle_class"    else ""
                    row["vehicle_category"] = name if section=="vehicle_category" else ""
                    row["fuel"]             = name if section=="fuel"             else ""
                    row["norms"]            = name if section=="norms"            else ""
                    row["registrations_count"] = count
                    rows.append(row)
                    added += 1

            if added == 0:
                rows.append({**base,"breakdown_type":"total","maker":"",
                    "vehicle_class":"","vehicle_category":"","fuel":"","norms":"",
                    "registrations_count":None})

    return rows

# ── Main loop ─────────────────────────────────────────────────────────────────
def run(driver, state_filter, months, progress):
    all_rows = []
    # Set type and wait for state dropdown to populate (retry up to 30s)
    # Discover dynamic JSF element IDs, then set type dropdown
    _ID_CACHE.clear()  # reset cache for fresh discovery each run
    ids = _discover_ids(driver)
    type_id  = ids['type_id']
    state_id = ids['state_id']
    rto_id   = ids['rto_id']

    # Try both 'A' (all) and 'Y' (yearwise) — site uses different values
    for type_val in ['A', 'Y']:
        pf_set(driver, type_id, type_val, wait=3)
        state_opts = []
        for _attempt in range(15):
            state_opts = get_opts(driver, state_id)
            if state_opts:
                break
            log.info(f"  Waiting for state dropdown (type={type_val}, attempt {_attempt+1})...")
            time.sleep(2)
            wait_jquery(driver)
        if state_opts:
            log.info(f"  Type value '{type_val}' worked — {len(state_opts)} states loaded")
            break

    log.info(f"States: {len(state_opts)}")
    if not state_opts:
        log.error("State dropdown empty — reloading page and retrying with fresh ID discovery")
        driver.get(BASE_URL)
        time.sleep(WAIT_PAGE); wait_jquery(driver)
        _ID_CACHE.clear()
        ids = _discover_ids(driver)
        type_id  = ids['type_id']
        state_id = ids['state_id']
        rto_id   = ids['rto_id']
        for type_val in ['A', 'Y']:
            pf_set(driver, type_id, type_val, wait=5)
            time.sleep(5); wait_jquery(driver)
            state_opts = get_opts(driver, state_id)
            if state_opts:
                break
        log.info(f"States after reload: {len(state_opts)}")

    if state_filter:
        sf = [s.upper() for s in state_filter]
        state_opts = [(v,t) for v,t in state_opts if v.upper() in sf]
        log.info(f"Filtered to: {[v for v,_ in state_opts]}")

    for (sc, sn) in state_opts:
        log.info(f"\n{'='*65}\nSTATE: {sn} (code={sc})\n{'='*65}")
        pf_set(driver, state_id, sc, wait=WAIT_STATE)
        wait_jquery(driver)

        rto_opts = get_opts(driver, rto_id)
        log.info(f"  RTOs: {len(rto_opts)}")

        for (rc, rl) in rto_opts:
            rk = f"{sc}::{rc}"
            if rk in progress["completed"]:
                log.info(f"  [SKIP] {rl}"); continue

            log.info(f"\n  RTO: {rl} (code={rc})")
            for attempt in range(RETRIES):
                try:
                    pf_set(driver, state_id, sc, wait=3); wait_jquery(driver)
                    pf_set(driver, rto_id, rc, wait=4); wait_jquery(driver)
                    click_refresh(driver,wait=WAIT_RTO)

                    rto_rows = scrape_rto(driver,sc,sn,rc,rl,months)
                    all_rows.extend(rto_rows)
                    append_raw({"rto_key":rk,"rows":rto_rows})
                    progress["completed"].append(rk)
                    save_json(PROGRESS_FILE, progress)
                    log.info(f"  DONE {rl} -> {len(rto_rows)} rows")
                    break
                except KeyboardInterrupt: raise
                except Exception as e:
                    log.warning(f"  Attempt {attempt+1}/{RETRIES}: {e}")
                    if attempt == RETRIES-1:
                        progress["failed"].append(rk)
                        save_json(PROGRESS_FILE, progress)
                    else:
                        time.sleep(6)
                        try:
                            driver.get(BASE_URL); time.sleep(WAIT_PAGE)
                            _ID_CACHE.clear(); _discover_ids(driver)
                            pf_set(driver, _ID_CACHE.get('type_id','j_idt30_input'), 'A', wait=2)
                        except Exception: pass
            time.sleep(3)

        if all_rows: save_excel(all_rows)
        time.sleep(5)

    return all_rows

# ── Output ────────────────────────────────────────────────────────────────────
COLUMNS = ["time_period","month","year","state_code","state_name","rto_code",
           "rto_name","breakdown_type","maker","vehicle_class","vehicle_category",
           "fuel","norms","registrations_count"]

def save_excel(rows):
    if not rows: return
    df = pd.DataFrame(rows)
    for c in COLUMNS:
        if c not in df.columns: df[c]=""
    df = df[COLUMNS]
    df = df.assign(registrations_count=pd.to_numeric(df["registrations_count"],errors="coerce"))

    xlsx = OUT_DIR/"vahan_registrations.xlsx"
    with pd.ExcelWriter(xlsx,engine="openpyxl") as w:
        df.to_excel(w,sheet_name="Registrations",index=False)
        ws = w.sheets["Registrations"]
        ws.freeze_panes = "A2"
        widths = {"A":14,"B":8,"C":6,"D":12,"E":26,"F":10,"G":42,
                  "H":18,"I":32,"J":22,"K":22,"L":20,"M":20,"N":20}
        from openpyxl.styles import Font, PatternFill, Alignment
        fill = PatternFill("solid",fgColor="1F4E79")
        for cell in ws[1]:
            cell.font = Font(bold=True,color="FFFFFF",size=10)
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")
        for col,w_val in widths.items():
            ws.column_dimensions[col].width = w_val
    log.info(f"Saved {xlsx} ({len(df):,} rows)")
    df.to_csv(OUT_DIR/"vahan_registrations.csv",index=False,encoding="utf-8-sig")
    log.info(f"Saved CSV ({len(df):,} rows)")

# ── Entry point ───────────────────────────────────────────────────────────────
def main():
    p = argparse.ArgumentParser(description="VAHAN Scraper v5")
    p.add_argument("--browser",choices=["chrome","firefox"],default="chrome")
    p.add_argument("--states",nargs="*",default=None)
    p.add_argument("--headed",action="store_true")
    p.add_argument("--reset",action="store_true",help="Clear progress and restart")
    args = p.parse_args()

    if args.reset and PROGRESS_FILE.exists():
        PROGRESS_FILE.unlink(); log.info("Progress cleared.")

    months   = months_list()
    progress = load_json(PROGRESS_FILE,{"completed":[],"failed":[]})

    log.info("="*65)
    log.info("VAHAN Scraper v5")
    log.info(f"Browser : {args.browser} ({'headed' if args.headed else 'headless'})")
    log.info(f"Period  : {months[0]} to {months[-1]} ({len(months)} months)")
    log.info(f"States  : {args.states or 'ALL'}")
    log.info(f"Done    : {len(progress['completed'])} RTOs already scraped")
    log.info("="*65)

    driver = get_driver(args.browser, headed=args.headed)
    all_rows = []
    try:
        driver.get(BASE_URL)
        time.sleep(WAIT_PAGE); wait_jquery(driver)
        log.info(f"Page loaded: {driver.title}")
        all_rows = run(driver, args.states, months, progress)
    except KeyboardInterrupt:
        log.info("Interrupted — saving progress...")
    except Exception as e:
        log.exception(f"Fatal: {e}")
    finally:
        driver.quit()
        log.info("Browser closed.")

    if all_rows:
        save_excel(all_rows)
    else:
        raw = load_json(RAW_FILE,[])
        recovered = [r for e in raw for r in e.get("rows",[])]
        if recovered:
            log.info(f"Recovering {len(recovered):,} rows from raw file")
            save_excel(recovered)

    log.info(f"\nOutput  : {OUT_DIR.resolve()}")
    log.info(f"Done    : {len(progress['completed'])} RTOs")
    log.info(f"Failed  : {len(progress['failed'])} RTOs")
    if progress["failed"]: log.info(f"Failed  : {progress['failed']}")

    _run_postprocess(args)


# ═══════════════════════════════════════════════════════════════════════════
# SECTION 2 — FORMATTER (Wide Excel)
# ═══════════════════════════════════════════════════════════════════════════

MONTH_ORDER = ['JAN','FEB','MAR','APR','MAY','JUN','JUL','AUG','SEP','OCT','NOV','DEC']
ID_COLS = ['Period','Month','Year','State Code','State','RTO Code','RTO Name']

BREAKDOWNS = [
    ('fuel',             'fuel',             'BY FUEL TYPE'),
    ('vehicle_category', 'vehicle_category', 'BY VEHICLE CATEGORY'),
    ('vehicle_class',    'vehicle_class',    'BY VEHICLE CLASS'),
    ('norms',            'norms',            'BY EMISSION NORMS'),
    ('maker',            'maker',            'BY MAKER / BRAND'),
]
GROUP_LABELS = {
    'info':'IDENTIFICATION','fuel':'BY FUEL TYPE','vehicle_category':'BY VEHICLE CATEGORY',
    'vehicle_class':'BY VEHICLE CLASS','norms':'BY EMISSION NORMS',
    'maker':'BY MAKER / BRAND','total':'TOTAL',
}
GROUP_STYLE = {
    'info':             ('1F3864','D6E4F7'),
    'fuel':             ('833C00','FCE4D6'),
    'vehicle_category': ('375623','E2EFDA'),
    'vehicle_class':    ('4B2F8A','EAE0F7'),
    'norms':            ('1F4E79','DDEBF7'),
    'maker':            ('5C3317','F2E0D0'),
    'total':            ('404040','C6EFCE'),
}

# ── Style helpers ─────────────────────────────────────────────────────────────
def hf(sz=10):              return Font(bold=True,color='FFFFFF',name='Calibri',size=sz)
def df_(sz=9,bold=False):   return Font(name='Calibri',size=sz,bold=bold)
def fl(c):                  return PatternFill('solid',fgColor=c)
def bd():
    s=Side(style='thin',color='C0C0C0')
    return Border(left=s,right=s,top=s,bottom=s)
def ctr(): return Alignment(horizontal='center',vertical='center',wrap_text=False)
def lft(): return Alignment(horizontal='left',  vertical='center',wrap_text=False)

def auto_w(ws,min_w=6,max_w=28):
    for col in ws.columns:
        w=max(min_w,min(max_w,max((len(str(c.value)) if c.value is not None else 0) for c in col)+2))
        ws.column_dimensions[get_column_letter(col[0].column)].width=w

def clean_rto(name):
    m=re.match(r'^(.+?)\s*-\s*([A-Z]+\d+)\s*\(.*\)$',str(name).strip())
    return f"{m.group(2)} \u2013 {m.group(1).strip().title()}" if m else str(name).strip()

def clean_state(name):
    m=re.match(r'^(.+?)\s*\(',str(name).strip())
    return m.group(1).strip() if m else str(name).strip()

# ── Core builder ──────────────────────────────────────────────────────────────
def build_wide_sheet(df: pd.DataFrame, output_path: Path):
    df = df.copy()

    # Base grid: one row per RTO × month
    base = (df[['time_period','month','year','state_code','state_clean','rto_code','rto_clean']]
            .drop_duplicates()
            .copy()
            .rename(columns={'time_period':'Period','month':'Month','year':'Year',
                             'state_code':'State Code','state_clean':'State',
                             'rto_code':'RTO Code','rto_clean':'RTO Name'}))
    base = base.assign(_mn=base['Month'].map({m:i for i,m in enumerate(MONTH_ORDER)}))
    base = base.sort_values(['RTO Code','Year','_mn']).drop(columns='_mn').reset_index(drop=True)

    # Pivot each breakdown and merge onto base
    group_cols: dict[str, list] = {}
    wide = base.copy()

    for bt, cat_col, _ in BREAKDOWNS:
        sub = df[df['breakdown_type'] == bt].copy()
        if sub.empty:
            continue
        # Drop rows where the category value is purely numeric (scraper artefact —
        # these are RTO codes that leaked into the category column).
        # Note: vehicle_class entries start with "-" (e.g. "-Motor Car") so we
        # only drop values that are entirely digits, not ones starting with "-".
        sub = sub[~sub[cat_col].astype(str).str.strip()
                  .str.fullmatch(r'[0-9]+').fillna(False)].copy()
        if sub.empty:
            continue
        piv = (sub.pivot_table(
                    index=['time_period','rto_code'],
                    columns=cat_col,
                    values='registrations_count',
                    aggfunc='sum',
                    fill_value=0)
               .reset_index())
        piv.columns.name = None
        piv = piv.rename(columns={'time_period':'Period','rto_code':'RTO Code'})
        cat_names = [c for c in piv.columns if c not in ('Period','RTO Code')]
        group_cols[bt] = cat_names
        wide = wide.merge(piv, on=['Period','RTO Code'], how='left')

    # Fill NaN → 0 for all category columns
    all_cat_cols = [c for cols in group_cols.values() for c in cols]
    # Fix all numeric columns in one shot using assign — avoids pandas CoW warnings
    num_updates = {c: pd.to_numeric(wide[c], errors='coerce').fillna(0).astype(int)
                   for c in all_cat_cols if c in wide.columns}
    wide = wide.assign(**num_updates)

    # TOTAL = sum of fuel cols (or vehicle_category as fallback)
    total_src = [c for c in (group_cols.get('fuel') or group_cols.get('vehicle_category') or [])
                 if c in wide.columns]
    wide = wide.assign(**{'TOTAL REGISTRATIONS': wide[total_src].sum(axis=1) if total_src else 0})
    group_cols['total'] = ['TOTAL REGISTRATIONS']

    # Build final column list — only include cols that actually exist in wide
    final_cols = [c for c in ID_COLS if c in wide.columns]
    for bt, _, _ in BREAKDOWNS:
        final_cols += [c for c in group_cols.get(bt, []) if c in wide.columns]
    final_cols += ['TOTAL REGISTRATIONS']

    # Map each column to its group
    col_group = {c: 'info' for c in ID_COLS}
    for bt, _, _ in BREAKDOWNS:
        for c in group_cols.get(bt, []):
            col_group[c] = bt
    col_group['TOTAL REGISTRATIONS'] = 'total'

    # Safety: use reindex to guarantee all final_cols exist (fill missing with 0)
    wide = wide.reindex(columns=final_cols, fill_value=0)

    log.info(f"  Wide table: {len(wide)} rows x {len(final_cols)} columns")
    for bt, _, label in BREAKDOWNS:
        n = len([c for c in group_cols.get(bt, []) if c in wide.columns])
        if n: log.info(f"    {label:28}: {n} cols")

    # ── Write Excel ────────────────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = 'Vehicle Registration Data'

    # Find column index ranges per group
    grp_ranges: dict[str, tuple] = {}
    for ci, col in enumerate(final_cols, 1):
        grp = col_group.get(col, 'info')
        grp_ranges[grp] = (grp_ranges[grp][0], ci) if grp in grp_ranges else (ci, ci)

    # Row 1: merged group header cells
    for grp, (s, e) in grp_ranges.items():
        bg, _ = GROUP_STYLE[grp]
        sl, el = get_column_letter(s), get_column_letter(e)
        if s != e:
            ws.merge_cells(f'{sl}1:{el}1')
        cell = ws.cell(1, s, GROUP_LABELS.get(grp, grp.upper()))
        cell.font=hf(10); cell.fill=fl(bg); cell.alignment=ctr(); cell.border=bd()
        for ci2 in range(s+1, e+1):
            c=ws.cell(1,ci2); c.fill=fl(bg); c.border=bd()
    ws.row_dimensions[1].height = 22

    # Row 2: column name headers
    for ci, col in enumerate(final_cols, 1):
        grp = col_group.get(col, 'info')
        bg, _ = GROUP_STYLE[grp]
        c = ws.cell(2, ci, col)
        c.font=hf(9); c.fill=fl(bg); c.alignment=ctr(); c.border=bd()
    ws.row_dimensions[2].height = 44

    # Data rows
    data = wide[final_cols].values.tolist()
    for ri, row_vals in enumerate(data):
        er = ri + 3
        alt = ri % 2 == 0
        for ci, (col, val) in enumerate(zip(final_cols, row_vals), 1):
            grp = col_group.get(col, 'info')
            _, alt_bg = GROUP_STYLE[grp]
            # Convert numpy int64 to plain int for openpyxl
            if hasattr(val, 'item'):
                val = val.item()
            c = ws.cell(er, ci, val)
            c.border = bd()
            if grp == 'info':
                c.font = df_(9, bold=(col in ('RTO Code','RTO Name')))
                c.alignment = lft()
                if alt: c.fill = fl(alt_bg)
            elif col == 'TOTAL REGISTRATIONS':
                c.font = df_(9, bold=True); c.fill = fl(alt_bg)
                c.alignment = ctr(); c.number_format = '#,##0'
            else:
                c.font = df_(9); c.alignment = ctr(); c.number_format = '#,##0'
                if alt: c.fill = fl(alt_bg)
        ws.row_dimensions[er].height = 14

    ws.freeze_panes = ws.cell(3, len(ID_COLS) + 1)
    auto_w(ws)
    for col_letter, w in zip('ABCDEFG', [10, 7, 7, 12, 12, 10, 34]):
        ws.column_dimensions[col_letter].width = w
    ws.sheet_properties.tabColor = '1F3864'
    wb.save(output_path)
    log.info(f"Saved: {output_path}  ({round(output_path.stat().st_size/1024)} KB)")

# ── Public API ────────────────────────────────────────────────────────────────
def format_data(input_path, output_path, state_filter=None):
    log.info(f"Loading: {input_path}")
    df = pd.read_excel(input_path)
    df = df.copy()
    df['registrations_count'] = pd.to_numeric(df['registrations_count'], errors='coerce').fillna(0).astype(int)

    if state_filter:
        sf = [s.upper() for s in state_filter]
        df = df[df['state_code'].str.upper().isin(sf)].copy()
        log.info(f"Filtered to: {sf} -> {len(df):,} rows")

    if df.empty:
        log.error("No data after filtering.")
        return None

    df['rto_clean']   = df['rto_name'].apply(clean_rto)
    df['state_clean'] = df['state_name'].apply(clean_state)

    log.info(f"States : {', '.join(sorted(df['state_clean'].unique()))}")
    log.info(f"RTOs   : {df['rto_clean'].nunique()}  |  Rows: {len(df):,}")
    build_wide_sheet(df, Path(output_path))
    return output_path

# ── CLI ───────────────────────────────────────────────────────────────────────
def main():
    p = argparse.ArgumentParser(description='VAHAN Data Formatter')
    p.add_argument('--input',  default=None)
    p.add_argument('--output', default=None)
    p.add_argument('--states', nargs='*', default=None)
    args = p.parse_args()

    data_dir = Path('vahan_data')
    input_path = Path(args.input) if args.input else data_dir / 'vahan_registrations.xlsx'
    if not input_path.exists():
        log.error(f"Not found: {input_path}"); sys.exit(1)

    if args.output:
        output_path = Path(args.output)
    else:
        if args.states:
            codes = '_'.join(sorted(s.upper() for s in args.states))
        else:
            try:
                codes = '_'.join(sorted(
                    pd.read_excel(input_path, usecols=['state_code'])['state_code'].unique()))
            except Exception:
                codes = 'ALL'
        output_path = data_dir / f"VAHAN_{codes}_{datetime.now().strftime('%Y%m%d')}_formatted.xlsx"

    data_dir.mkdir(exist_ok=True)
    format_data(input_path, output_path, args.states)


# ═══════════════════════════════════════════════════════════════════════════
# SECTION 3 — BIGQUERY SIDE-BY-SIDE EXPORTER
# ═══════════════════════════════════════════════════════════════════════════

GROUP_COLORS = {
    'id':               ('1F3864', 'D6E4F7'),
    'fuel':             ('833C00', 'FCE4D6'),
    'vehicle_category': ('375623', 'E2EFDA'),
    'vehicle_class':    ('4B2F8A', 'EAE0F7'),
    'norms':            ('1F4E79', 'DDEBF7'),
    'maker':            ('5C3317', 'F2E0D0'),
}

MASTER_FUEL_MAP = {
    'PETROL':'Petrol','DIESEL':'Diesel','CNG ONLY':'CNG','PETROL/CNG':'CNG',
    'ELECTRIC(BOV)':'Electric','PURE EV':'Electric','STRONG HYBRID EV':'Strong Hybrid EV',
    'PETROL/HYBRID':'Petrol - Mild Hybrid','DIESEL/HYBRID':'Diesel - Mild Hybrid',
    'PETROL(E20)/HYBRID':'Petrol - Mild Hybrid','PETROL/ETHANOL':'Petrol - Ethanol',
    'FUEL CELL HYDROGEN':'Hydrogen','LPG ONLY':'LPG','PETROL/LPG':'LPG',
}

OEM_MAP = {
    "MARUTI SUZUKI INDIA LTD":"Maruti","MARUTI UDYOG LTD":"Maruti",
    "HYUNDAI MOTOR INDIA LTD":"Hyundai","TATA MOTORS LTD":"Tata",
    "TATA MOTORS PASSENGER VEHICLES LTD":"Tata","TATA PASSENGER ELECTRIC MOBILITY LTD":"Tata",
    "MAHINDRA & MAHINDRA LIMITED":"Mahindra","MAHINDRA & MAHINDRA LIMITED (SWARAJ DIVISION)":"Swaraj",
    "MAHINDRA LAST MILE MOBILITY LTD":"Mahindra-LMM",
    "TOYOTA KIRLOSKAR MOTOR PVT LTD":"Toyota","HONDA CARS INDIA LTD":"Honda",
    "HONDA MOTORCYCLE AND SCOOTER INDIA (P) LTD":"Honda","KIA INDIA PRIVATE LIMITED":"Kia",
    "SKODA AUTO VOLKSWAGEN INDIA PVT LTD":"Volkswagen","RENAULT INDIA PVT LTD":"Renault",
    "NISSAN MOTOR INDIA PVT LTD":"Nissan","JSW MG MOTOR INDIA PVT LTD":"MG",
    "BMW INDIA PVT LTD":"BMW","MERCEDES-BENZ INDIA PVT LTD":"Mercedes-Benz",
    "MERCEDES -BENZ AG":"Mercedes-Benz","AUDI AG":"Audi","PORSCHE AG GERMANY":"Porsche",
    "JAGUAR LAND ROVER INDIA LIMITED":"JLR","FORD INDIA PVT LTD":"Ford",
    "ISUZU MOTORS INDIA PVT LTD":"Isuzu","SML ISUZU LTD":"SML Isuzu",
    "HERO MOTOCORP LTD":"Hero","HERO HONDA MOTORS LTD":"Hero",
    "HERO ELECTRIC VEHICLES PVT. LTD":"Hero Electric","BAJAJ AUTO LTD":"Bajaj",
    "TVS MOTOR COMPANY LTD":"TVS","ROYAL-ENFIELD (UNIT OF EICHER LTD)":"Royal Enfield",
    "INDIA YAMAHA MOTOR PVT LTD":"Yamaha","SUZUKI MOTORCYCLE INDIA PVT LTD":"Suzuki",
    "INDIA KAWASAKI MOTORS PVT LTD":"Kawasaki","ATHER ENERGY LTD":"Ather",
    "OLA ELECTRIC TECHNOLOGIES PVT LTD":"Ola Electric","REVOLT INTELLICORP PVT LTD":"Revolt",
    "ASHOK LEYLAND LTD":"Ashok Leyland","VE COMMERCIAL VEHICLES LTD":"Volvo Eicher",
    "DAIMLER INDIA COMMERCIAL VEHICLES PVT. LTD":"Daimler","VOLVO GROUP INDIA PVT LTD":"Volvo",
    "VOLVO AUTO INDIA PVT LTD":"Volvo","OLECTRA GREENTECH LTD":"Olectra",
    "JBM AUTO LIMITED":"JBM","FORCE MOTORS LIMITED":"Force Motors",
    "PIAGGIO VEHICLES PVT LTD":"Piaggio","ATUL AUTO LTD":"Atul Auto",
    "EULER MOTORS PVT LTD":"Euler Motors","INTERNATIONAL TRACTORS LIMITED":"Sonalika",
    "JOHN DEERE INDIA PVT LTD(TRACTOR DEVISION)":"John Deere",
    "ESCORTS KUBOTA LIMITED (AGRI MACHINERY GROUP)":"Escorts Kubota",
    "CNH INDUSTRIAL (INDIA) PVT LTD":"CNH","TAFE LIMITED":"TAFE",
    "JCB INDIA LIMITED":"JCB","KOMATSU INDIA PRIVATE LIMITED":"Komatsu",
    "CATERPILLAR INDIA PRIVATE LIMITED":"Caterpillar","SANY HEAVY INDUSTRY INDIA PVT LTD":"Sany",
    "BYD INDIA PRIVATE LIMITED":"BYD","AJAX ENGINEERING LTD":"Ajax",
    "ACTION CONSTRUCTION EQUIPMENT LTD.":"ACE","TESLA INDIA MOTORS AND ENERGY PVT LTD":"Tesla",
    "GREAVES ELECTRIC MOBILITY PVT LTD":"Greaves Electric","MAHINDRA ELECTRIC AUTOMOBILE LTD":"Mahindra",
    "MAHINDRA ELECTRIC MOBILITY LIMITED":"Mahindra",
}

def get_oem(maker):
    if not maker or pd.isna(maker): return ''
    m = str(maker).strip()
    if m in OEM_MAP: return OEM_MAP[m]
    ml = m.upper()
    for k,v in OEM_MAP.items():
        if k.upper() in ml: return v
    words = re.sub(r'\(.*?\)', '', m).strip().split()
    return ' '.join(words[:2]).title() if words else m

def get_master_fuel(f):
    if not f or pd.isna(f): return ''
    return MASTER_FUEL_MAP.get(str(f).strip().upper(), str(f).strip().title())

def get_fy(tp):
    y, m = int(str(tp)[:4]), int(str(tp)[5:7])
    return f"F{str(y+1 if m>=4 else y)[2:]}"

def get_quarter(tp):
    y, m = int(str(tp)[:4]), int(str(tp)[5:7])
    fy = y+1 if m>=4 else y
    q = ((m-4)%12)//3+1
    return f"F{str(fy)[2:]}Q{q}"

def clean_rto(name):
    m = re.match(r'^(.+?)\s*-\s*([A-Z]+\d+)\s*\(.*\)$', str(name).strip())
    return f"{m.group(2)} \u2013 {m.group(1).strip().title()}" if m else str(name).strip()

def clean_state(name):
    m = re.match(r'^(.+?)\s*\(', str(name).strip())
    return m.group(1).strip() if m else str(name).strip()

def build_sidebyside(raw_xlsx: Path, output_xlsx: Path, state_filter=None):
    print(f"Loading: {raw_xlsx}")
    df = pd.read_excel(raw_xlsx)
    df['registrations_count'] = pd.to_numeric(df['registrations_count'], errors='coerce').fillna(0).astype(int)

    if state_filter:
        df = df[df['state_code'].str.upper().isin([s.upper() for s in state_filter])].copy()

    df = df[df['registrations_count'] > 0].copy()
    df = df[df['breakdown_type'] != 'total'].copy()
    df['rto_clean']   = df['rto_name'].apply(clean_rto)
    df['state_clean'] = df['state_name'].apply(clean_state)

    created_dt = datetime.now().strftime('%Y-%m-%d %H:%M:%S UTC')

    # ── Build each breakdown sub-table ────────────────────────────────────────
    ID = ['financial_year','l_quarter','month','state_code','state_name','rto_code','rto_name']

    def make_base(row):
        tp = str(row['time_period'])
        return {
            'financial_year': get_fy(tp),
            'l_quarter':      get_quarter(tp),
            'month':          f"{tp}-01",
            'state_code':     row['state_code'],
            'state_name':     row['state_clean'],
            'rto_code':       row['rto_code'],
            'rto_name':       row['rto_clean'],
        }

    # Fuel
    fuel_rows = []
    for _, r in df[df['breakdown_type']=='fuel'].iterrows():
        b = make_base(r)
        f = str(r.get('fuel','') or '').strip()
        b['breakdown_type_fuel']  = 'fuel'
        b['master_fuel']          = get_master_fuel(f)
        b['fuel']                 = f
        b['registrations_fuel']   = int(r['registrations_count'])
        fuel_rows.append(b)
    fuel_df = pd.DataFrame(fuel_rows)

    # Vehicle Category
    vc_rows = []
    for _, r in df[df['breakdown_type']=='vehicle_category'].iterrows():
        vc_rows.append({
            'breakdown_type_vehicle_category': 'vehicle_category',
            'vehicle_category':                str(r.get('vehicle_category','') or '').strip(),
            'registrations_vehicle_category':  int(r['registrations_count']),
        })
    vc_df = pd.DataFrame(vc_rows)

    # Vehicle Class
    vcl_rows = []
    for _, r in df[df['breakdown_type']=='vehicle_class'].iterrows():
        vcl_rows.append({
            'breakdown_type_vehicle_class': 'vehicle_class',
            'vehicle_class':                str(r.get('vehicle_class','') or '').strip(),
            'registrations_vehicle_class':  int(r['registrations_count']),
        })
    vcl_df = pd.DataFrame(vcl_rows)

    # Norms
    norms_rows = []
    for _, r in df[df['breakdown_type']=='norms'].iterrows():
        norms_rows.append({
            'breakdown_type_norms': 'norms',
            'norms':                str(r.get('norms','') or '').strip(),
            'registrations_norms':  int(r['registrations_count']),
        })
    norms_df = pd.DataFrame(norms_rows)

    # Maker
    maker_rows = []
    for _, r in df[df['breakdown_type']=='maker'].iterrows():
        mk = str(r.get('maker','') or '').strip()
        maker_rows.append({
            'breakdown_type_maker': 'maker',
            'maker':                mk,
            'oem':                  get_oem(mk),
            'registrations_maker':  int(r['registrations_count']),
        })
    maker_df = pd.DataFrame(maker_rows)

    # ── Combine side by side (align by row index) ─────────────────────────────
    combined = pd.concat([
        fuel_df.reset_index(drop=True),
        vc_df.reset_index(drop=True),
        vcl_df.reset_index(drop=True),
        norms_df.reset_index(drop=True),
        maker_df.reset_index(drop=True),
    ], axis=1)

    combined['mdp_created_dt'] = created_dt

    print(f"Shape: {combined.shape}")
    print(f"Columns ({len(combined.columns)}): {combined.columns.tolist()[:10]}...")

    # ── Write to Excel ─────────────────────────────────────────────────────────
    COLS_ORDERED = (
        ID +
        ['breakdown_type_fuel','master_fuel','fuel','registrations_fuel'] +
        ['breakdown_type_vehicle_category','vehicle_category','registrations_vehicle_category'] +
        ['breakdown_type_vehicle_class','vehicle_class','registrations_vehicle_class'] +
        ['breakdown_type_norms','norms','registrations_norms'] +
        ['breakdown_type_maker','maker','oem','registrations_maker'] +
        ['mdp_created_dt']
    )
    # Only keep cols that exist
    COLS_ORDERED = [c for c in COLS_ORDERED if c in combined.columns]
    combined = combined.reindex(columns=COLS_ORDERED)

    # Column display names
    COL_DISPLAY = {
        'financial_year':'financial_year','l_quarter':'l_quarter','month':'month',
        'state_code':'state_code','state_name':'state_name',
        'rto_code':'rto_code','rto_name':'rto_name',
        'breakdown_type_fuel':'breakdown_type','master_fuel':'master_fuel',
        'fuel':'fuel','registrations_fuel':'registrations',
        'breakdown_type_vehicle_category':'breakdown_type',
        'vehicle_category':'vehicle_category','registrations_vehicle_category':'registrations',
        'breakdown_type_vehicle_class':'breakdown_type',
        'vehicle_class':'vehicle_class','registrations_vehicle_class':'registrations',
        'breakdown_type_norms':'breakdown_type',
        'norms':'norms','registrations_norms':'registrations',
        'breakdown_type_maker':'breakdown_type',
        'maker':'maker','oem':'oem','registrations_maker':'registrations',
        'mdp_created_dt':'mdp_created_dt',
    }

    # Group membership for coloring
    COL_GROUP = {}
    for c in ID: COL_GROUP[c] = 'id'
    for c in ['breakdown_type_fuel','master_fuel','fuel','registrations_fuel']:
        COL_GROUP[c] = 'fuel'
    for c in ['breakdown_type_vehicle_category','vehicle_category','registrations_vehicle_category']:
        COL_GROUP[c] = 'vehicle_category'
    for c in ['breakdown_type_vehicle_class','vehicle_class','registrations_vehicle_class']:
        COL_GROUP[c] = 'vehicle_class'
    for c in ['breakdown_type_norms','norms','registrations_norms']:
        COL_GROUP[c] = 'norms'
    for c in ['breakdown_type_maker','maker','oem','registrations_maker']:
        COL_GROUP[c] = 'maker'
    COL_GROUP['mdp_created_dt'] = 'id'

    COL_W = {
        'financial_year':13,'l_quarter':11,'month':12,'state_code':10,'state_name':12,
        'rto_code':9,'rto_name':30,
        'breakdown_type_fuel':14,'master_fuel':20,'fuel':18,'registrations_fuel':13,
        'breakdown_type_vehicle_category':15,'vehicle_category':24,'registrations_vehicle_category':13,
        'breakdown_type_vehicle_class':15,'vehicle_class':28,'registrations_vehicle_class':13,
        'breakdown_type_norms':13,'norms':22,'registrations_norms':13,
        'breakdown_type_maker':13,'maker':36,'oem':16,'registrations_maker':13,
        'mdp_created_dt':22,
    }

    # Group header row 1
    GROUP_SPANS = [
        ('IDENTIFICATION',   'id',               [c for c in COLS_ORDERED if COL_GROUP.get(c)=='id' and c!='mdp_created_dt']),
        ('FUEL',             'fuel',             [c for c in COLS_ORDERED if COL_GROUP.get(c)=='fuel']),
        ('VEHICLE CATEGORY', 'vehicle_category', [c for c in COLS_ORDERED if COL_GROUP.get(c)=='vehicle_category']),
        ('VEHICLE CLASS',    'vehicle_class',    [c for c in COLS_ORDERED if COL_GROUP.get(c)=='vehicle_class']),
        ('EMISSION NORMS',   'norms',            [c for c in COLS_ORDERED if COL_GROUP.get(c)=='norms']),
        ('MAKER / BRAND',    'maker',            [c for c in COLS_ORDERED if COL_GROUP.get(c)=='maker']),
        ('METADATA',         'id',               ['mdp_created_dt'] if 'mdp_created_dt' in COLS_ORDERED else []),
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = 'VAHAN Data'

    # Row 1: group headers
    col_pos = {c: i+1 for i, c in enumerate(COLS_ORDERED)}
    for label, grp, cols in GROUP_SPANS:
        if not cols: continue
        s = col_pos[cols[0]]
        e = col_pos[cols[-1]]
        bg, _ = GROUP_COLORS.get(grp, ('1F3864','FFFFFF'))
        if s != e:
            ws.merge_cells(f'{get_column_letter(s)}1:{get_column_letter(e)}1')
        cell = ws.cell(1, s, label)
        cell.font = hf(10); cell.fill = fl(bg); cell.alignment = ctr(); cell.border = bd()
        for ci in range(s+1, e+1):
            c = ws.cell(1, ci); c.fill = fl(bg); c.border = bd()
    ws.row_dimensions[1].height = 20

    # Row 2: column headers
    for ci, col in enumerate(COLS_ORDERED, 1):
        grp = COL_GROUP.get(col, 'id')
        bg, _ = GROUP_COLORS.get(grp, ('1F3864','FFFFFF'))
        c = ws.cell(2, ci, COL_DISPLAY.get(col, col))
        c.font = hf(9); c.fill = fl(bg); c.alignment = ctr(); c.border = bd()
        ws.column_dimensions[get_column_letter(ci)].width = COL_W.get(col, 14)
    ws.row_dimensions[2].height = 38

    # Data rows
    for ri, row_vals in enumerate(combined[COLS_ORDERED].values.tolist(), 3):
        alt = ri % 2 == 0
        for ci, (col, val) in enumerate(zip(COLS_ORDERED, row_vals), 1):
            grp = COL_GROUP.get(col, 'id')
            _, alt_bg = GROUP_COLORS.get(grp, ('1F3864','F2F7FD'))
            if hasattr(val, 'item'): val = val.item()
            # Convert NaN to empty string
            if pd.isna(val) if not isinstance(val, str) else False: val = ''
            c = ws.cell(ri, ci, val)
            c.font = df_(9); c.border = bd()
            is_reg = col.startswith('registrations')
            c.alignment = ctr() if is_reg or col in ('rto_code','state_code','financial_year','l_quarter','month') else lft()
            if is_reg and isinstance(val, (int, float)) and val != '':
                c.number_format = '#,##0'
            if alt: c.fill = fl(alt_bg)
        ws.row_dimensions[ri].height = 13

    ws.freeze_panes = 'H3'
    ws.sheet_properties.tabColor = '1F3864'
    wb.save(output_xlsx)
    print(f"Saved: {output_xlsx}  ({round(Path(output_xlsx).stat().st_size/1024)} KB)")


# ═══════════════════════════════════════════════════════════════════════════
# POST-PROCESS ORCHESTRATOR
# ═══════════════════════════════════════════════════════════════════════════

def _run_postprocess(args):
    """Step 2 + Step 3 after scraping."""
    raw_xlsx = OUT_DIR / "vahan_registrations.xlsx"
    if not raw_xlsx.exists():
        log.warning(f"No raw file at {raw_xlsx} — skipping post-processing")
        return
    try:
        state_codes = args.states if args.states else None
        if state_codes:
            codes = "_".join(sorted(s.upper() for s in state_codes))
        else:
            codes = "_".join(sorted(
                pd.read_excel(raw_xlsx, usecols=["state_code"])["state_code"].unique().tolist()))
        ts = datetime.now().strftime("%Y%m%d")

        # Step 2: Wide pivot Excel
        wide_path = OUT_DIR / f"VAHAN_{codes}_{ts}_formatted.xlsx"
        log.info("="*65)
        log.info(f"Step 2: Wide Excel → {wide_path.name}")
        log.info("="*65)
        format_data(raw_xlsx, wide_path, state_filter=state_codes)
        log.info(f"Saved: {wide_path.resolve()}")

        # Step 3: BigQuery side-by-side Excel
        bq_path = OUT_DIR / f"VAHAN_{codes}_{ts}_bigquery.xlsx"
        log.info("="*65)
        log.info(f"Step 3: BigQuery Excel → {bq_path.name}")
        log.info("="*65)
        build_sidebyside(raw_xlsx, bq_path, state_filter=state_codes)
        log.info(f"Saved: {bq_path.resolve()}")

    except Exception as _fe:
        log.warning(f"Post-processing failed: {_fe}")
        import traceback; traceback.print_exc()


# ═══════════════════════════════════════════════════════════════════════════
# MAIN ENTRY POINT
# ═══════════════════════════════════════════════════════════════════════════

def main():
    p = argparse.ArgumentParser(description="VAHAN Pipeline — Scrape + Format + BigQuery")
    p.add_argument("--browser", choices=["chrome","firefox"], default="chrome")
    p.add_argument("--states",  nargs="*", default=None)
    p.add_argument("--headed",  action="store_true")
    p.add_argument("--reset",   action="store_true", help="Clear progress and restart")
    p.add_argument("--format-only", action="store_true",
                   help="Skip scraping; re-format existing vahan_registrations.xlsx")
    args = p.parse_args()
    args.format_only = getattr(args, 'format_only', False)

    if args.reset and PROGRESS_FILE.exists():
        PROGRESS_FILE.unlink(); log.info("Progress cleared.")

    months   = months_list()
    progress = load_json(PROGRESS_FILE, {"completed":[], "failed":[]})

    log.info("="*65)
    log.info("VAHAN Pipeline")
    log.info(f"Browser : {args.browser} ({'headed' if args.headed else 'headless'})")
    log.info(f"Period  : {months[0]} to {months[-1]} ({len(months)} months)")
    log.info(f"States  : {args.states or 'ALL'}")
    log.info(f"Done    : {len(progress['completed'])} RTOs already scraped")
    log.info("="*65)

    if args.format_only:
        log.info("--format-only: skipping scrape")
        _run_postprocess(args)
        return

    driver = get_driver(args.browser, headed=args.headed)
    all_rows = []
    try:
        driver.get(BASE_URL)
        time.sleep(WAIT_PAGE); wait_jquery(driver)
        log.info(f"Page loaded: {driver.title}")
        all_rows = run(driver, args.states, months, progress)
    except KeyboardInterrupt:
        log.info("Interrupted — saving progress...")
    except Exception as e:
        log.exception(f"Fatal: {e}")
    finally:
        driver.quit()
        log.info("Browser closed.")

    if all_rows:
        save_excel(all_rows)
    else:
        raw = load_json(RAW_FILE, [])
        recovered = [r for e in raw for r in e.get("rows", [])]
        if recovered:
            log.info(f"Recovering {len(recovered):,} rows from raw file")
            save_excel(recovered)

    log.info(f"\nOutput  : {OUT_DIR.resolve()}")
    log.info(f"Done    : {len(progress['completed'])} RTOs")
    log.info(f"Failed  : {len(progress['failed'])} RTOs")
    if progress["failed"]: log.info(f"Failed  : {progress['failed']}")

    _run_postprocess(args)


if __name__ == "__main__":
    main()
